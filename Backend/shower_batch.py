#!/usr/bin/env python3
"""Batch runner for shower programming jobs listed in the process workbook."""

from __future__ import annotations

# UNMARKED_SKETCH_AND_ORDER_PDF_V12: process-list mapping for unmarked pages and order-number PDF names.
# OOS_DIMENSION_RECONCILIATION_V93: reconcile A+W overall size with sketch edge size using proven DXF OOS geometry.
# DXF_SKETCH_ENVELOPE_RECONCILIATION_V111: reconcile near A+W irregular dimensions when sketch and source DXF envelope agree.
# DXF_GEOMETRY_DIMENSION_RECONCILIATION_V94: accept alternate sketch edge dimensions only when source DXF geometry proves them.
# LEGACY_XLS_FAST_CONVERSION_V115: reuse content-identical conversions and suppress slow Excel update/event work.

import argparse
import csv
import hashlib
import html
import math
import os
import re
import subprocess
import sys
import time
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook
from pypdf import PdfReader

import shower_programmer as programmer
import shower_cache
import shower_errors


DEFAULT_PROCESS_LIST = "Process List Per Machine.xlsx"
PROCESS_LIST_EXTENSIONS = {".xlsx", ".xml", ".rtf", ".xls"}
PROCESS_LIST_EXTENSION_PRIORITY = {".xlsx": 0, ".xml": 1, ".rtf": 2, ".xls": 3}
PROCESS_LIST_CACHE_NAMESPACE = "process_orders_v1"
PROCESS_LIST_CACHE_SCHEMA = 2
ProcessListProgress = Callable[[str, Path, str], None]
PDF_DIMENSION_MATCH_TOLERANCE = 0.20
OOS_DXF_OVERALL_TOLERANCE = 0.08
# When the sketch dimensions match the source DXF envelope exactly, allow a
# small A+W process-size variance for proven irregular/OOS geometry. This is
# intentionally separate from the normal PDF tolerance and from the stricter
# A+W-to-DXF overall check used by the original OOS rule.
OOS_PROCESS_DXF_VARIANCE_TOLERANCE = 0.25
OOS_DXF_EDGE_TOLERANCE = 0.35
OOS_MIN_SHIFT_INCHES = 1.0 / 32.0
OOS_MAX_SHIFT_INCHES = 2.0

GLASS_MATERIAL_PATTERN = re.compile(
    r'^\s*(?:\d+-\d+/\d+|\d+/\d+|0?\.\d+)\s*(?:"|IN(?:CH(?:ES)?)?)?\s+[A-Z].*$',
    re.IGNORECASE,
)


def upper_config_list(config: dict[str, object], key: str, default: list[str]) -> set[str]:
    rules = config.get("rules", {})
    if not isinstance(rules, dict):
        return {value.upper() for value in default}
    values = rules.get(key, default)
    if not isinstance(values, list):
        return {value.upper() for value in default}
    return {str(value).upper() for value in values}


@dataclass
class ProcessItem:
    item: int
    width_text: str = ""
    height_text: str = ""
    delivery_date: str = ""
    customer: str = ""
    processing: list[str] = field(default_factory=list)
    machine_hints: list[str] = field(default_factory=list)
    rows: list[int] = field(default_factory=list)

    def add_row(
        self,
        row_number: int,
        width_text: str,
        height_text: str,
        delivery_date: str,
        customer: str,
        processing: str,
        machine_hint: str,
    ) -> None:
        self.rows.append(row_number)
        if width_text and not self.width_text:
            self.width_text = width_text
        if height_text and not self.height_text:
            self.height_text = height_text
        if delivery_date and not self.delivery_date:
            self.delivery_date = delivery_date
        if customer and not self.customer:
            self.customer = customer
        if processing and processing not in self.processing:
            self.processing.append(processing)
        if machine_hint and machine_hint not in self.machine_hints:
            self.machine_hints.append(machine_hint)

    @property
    def processing_text(self) -> str:
        return " | ".join(self.processing)

    @property
    def machine_text(self) -> str:
        return " | ".join(self.machine_hints)

    @property
    def has_diamon_fusion(self) -> bool:
        return "DIAMON" in self.processing_text.upper() or "DIAMOND FUSION" in self.processing_text.upper()

    def desired_machine(self) -> str:
        text = self.machine_text.upper()
        if "WATER" in text or re.search(r"\bWJ\b", text):
            return "WJ"
        if "DENVER 1" in text:
            return "DENVER 1"
        if "DENVER 2" in text:
            return "DENVER 2"
        return ""

    def inferred_denver_machine(self, config: dict[str, object]) -> str:
        text = self.processing_text.upper()
        if not text:
            return ""
        fabrication_text = programmer.strip_non_fabrication_edge_text(text)
        door_keywords = upper_config_list(config, "door_keywords", ["DOOR", "HINGE", "PPH", "PULL", "HANDLE"])
        door_keywords.update(programmer.hinge_label_keywords(config))
        fabrication = process_list_fabrication_keywords(config)
        if (
            any(keyword in text for keyword in door_keywords)
            or programmer.has_hinge_label_text(text, config)
            or re.search(r"\b(?:A?V1E|A?GEN)\d{3}\b", text)
        ):
            return "DENVER 1"
        if any(keyword in fabrication_text for keyword in fabrication):
            return "DENVER 2"
        return ""

    def has_strong_waterjet_fabrication(self, config: dict[str, object]) -> bool:
        text = self.processing_text.upper()
        if not text:
            return False
        if re.search(r"\b(?:DOUBLE|TRIPLE)\s+NOTCH(?:ES)?\b", text):
            return True
        if re.search(r"\b[1-9]\d*\s+(?:EDGE\s+NOTCH(?:ES)?|CORNER\s+NOTCH(?:ES)?|NOTCHED\s+CORNERS?)\b", text):
            return True
        if re.search(r"\b(?:[1-9]\d*\s+)?(?:1/2\s+)?RADIUS\b", text):
            return True
        return False

    def text_blob(self) -> str:
        return " ".join(
            [
                self.width_text,
                self.height_text,
                self.delivery_date,
                self.customer,
                self.processing_text,
                self.machine_text,
            ]
        ).upper()

    def is_mirror(self, config: dict[str, object]) -> bool:
        keywords = upper_config_list(config, "mirror_keywords", ["MIRROR"])
        text = self.text_blob()
        return any(keyword in text for keyword in keywords)

    def has_mirror_fabrication(self, config: dict[str, object]) -> bool:
        keywords = upper_config_list(
            config,
            "mirror_fabrication_keywords",
            ["FAB", "GEN", "HOLE", "NOTCH", "CUTOUT", "CUT-OUT", "RADIUS"],
        )
        text = self.processing_text.upper()
        return any(keyword in text for keyword in keywords)


@dataclass
class ProcessOrder:
    aw_order: str
    job_name: str
    customer: str = ""
    items: dict[int, ProcessItem] = field(default_factory=dict)

    @property
    def item_numbers(self) -> list[int]:
        return sorted(self.items)

    @property
    def item_count(self) -> int:
        return len(self.items)

    @property
    def delivery_date(self) -> str:
        for item in self.items.values():
            if item.delivery_date:
                return item.delivery_date
        return ""

    def text_blob(self) -> str:
        return " ".join(
            [self.job_name, self.customer]
            + [item.text_blob() for item in self.items.values()]
        ).upper()

    def is_mirror(self, config: dict[str, object]) -> bool:
        keywords = upper_config_list(config, "mirror_keywords", ["MIRROR"])
        text = self.text_blob()
        return any(keyword in text for keyword in keywords)

    def has_mirror_fabrication(self, config: dict[str, object]) -> bool:
        return any(item.has_mirror_fabrication(config) for item in self.items.values())


@dataclass
class BatchJobResult:
    aw_order: str
    job_name: str
    customer: str
    items: str
    status: str
    input_pdf: Path | None = None
    output_pdf: Path | None = None
    report_path: Path | None = None
    delivery_date: str = ""
    issues: list[str] = field(default_factory=list)
    remake_items: list[int] | None = None


@dataclass
class BatchRunResult:
    results: list[BatchJobResult]
    text_report: Path
    csv_report: Path
    html_report: Path


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    return re.sub(r"\s+", " ", str(value)).strip()


def decode_text_file(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "utf-16le", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("latin-1", errors="replace")


def read_file_prefix(path: Path, size: int) -> bytes:
    with path.open("rb") as handle:
        return handle.read(size)


def parse_order_item(value: str) -> tuple[str, int] | None:
    match = re.search(r"\b(?P<order>\d{5,})\s*[-_]\s*(?P<item>\d+)\b", value)
    if not match:
        return None
    return match.group("order"), int(match.group("item"))


def process_list_files(path: Path) -> list[Path]:
    if path.is_dir():
        files = [
            candidate
            for candidate in path.iterdir()
            if is_process_list_file(candidate)
        ]
        return preferred_process_list_exports(files)
    if path.is_file():
        if not is_process_list_file(path):
            raise RuntimeError(
                f"Unsupported process-list file type: {path.name}. "
                f"Use one of: {process_list_extension_text()}."
            )
        return [path]
    raise FileNotFoundError(f"Process list path not found: {path}")


def preferred_process_list_exports(files: Iterable[Path]) -> list[Path]:
    preferred: dict[str, Path] = {}
    for path in sorted(files, key=lambda candidate: candidate.name.lower()):
        key = path.stem.lower()
        current = preferred.get(key)
        if current is None or process_list_extension_rank(path) < process_list_extension_rank(current):
            preferred[key] = path
    return sorted(preferred.values(), key=lambda candidate: candidate.name.lower())


def process_list_extension_rank(path: Path) -> int:
    return PROCESS_LIST_EXTENSION_PRIORITY.get(path.suffix.lower(), 99)


def is_process_list_file(path: Path) -> bool:
    return (
        path.is_file()
        and not path.name.startswith("~$")
        and path.suffix.lower() in PROCESS_LIST_EXTENSIONS
    )


def process_list_extension_text() -> str:
    return ", ".join(sorted(PROCESS_LIST_EXTENSIONS))


def load_process_orders(path: Path) -> list[ProcessOrder]:
    files = process_list_files(path)
    if not files:
        raise FileNotFoundError(
            f"No supported process lists ({process_list_extension_text()}) found in: {path}"
        )
    merged: dict[tuple[str, str], ProcessOrder] = {}
    errors: list[str] = []
    for process_list_path in files:
        try:
            process_orders = load_process_orders_from_file(process_list_path)
        except RuntimeError as exc:
            errors.append(f"{process_list_path.name}: {exc}")
            continue
        for order in process_orders:
            merge_process_order(merged, order)
    if not merged and errors:
        raise RuntimeError("No process-list orders could be loaded.\n" + "\n".join(errors))
    return sorted(merge_process_orders_by_aw(list(merged.values())), key=lambda order: (int(order.aw_order), order.job_name))


def load_process_orders_from_file(
    path: Path,
    progress_callback: ProcessListProgress | None = None,
) -> list[ProcessOrder]:
    path = Path(path).resolve()
    cached = shower_cache.load(PROCESS_LIST_CACHE_NAMESPACE, path)
    if isinstance(cached, dict) and cached.get("schema") == PROCESS_LIST_CACHE_SCHEMA:
        try:
            orders = process_orders_from_cache(cached.get("orders", []))
        except Exception:
            orders = []
        if orders:
            if progress_callback:
                progress_callback("cached", path, f"Reused {len(orders)} cached order(s)")
            return orders
    if progress_callback:
        progress_callback("loading", path, "Reading process-list rows")
    orders = load_process_orders_from_file_uncached(path, progress_callback)
    shower_cache.store(
        PROCESS_LIST_CACHE_NAMESPACE,
        path,
        {"schema": PROCESS_LIST_CACHE_SCHEMA, "orders": process_orders_to_cache(orders)},
    )
    if progress_callback:
        progress_callback("loaded", path, f"Loaded {len(orders)} order(s)")
    return orders


def load_process_orders_from_file_uncached(
    path: Path,
    progress_callback: ProcessListProgress | None = None,
) -> list[ProcessOrder]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return load_process_orders_from_workbook(path)
    if suffix == ".xml":
        rows = load_rows_from_spreadsheet_xml(path)
        if not rows:
            rows = load_rows_from_crystal_xml(path)
        return load_process_orders_from_rows(rows)
    if suffix == ".rtf":
        orders = load_process_orders_from_rows(load_rows_from_rtf(path))
        if orders:
            return orders
        return load_process_orders_from_rows(load_rows_from_crystal_rtf(path))
    if suffix == ".xls":
        return load_process_orders_from_legacy_xls(path, progress_callback)
    raise RuntimeError(
        f"Unsupported process-list file type: {path.name}. Use one of: {process_list_extension_text()}."
    )


def process_orders_to_cache(orders: list[ProcessOrder]) -> list[dict[str, object]]:
    return [
        {
            "aw_order": order.aw_order,
            "job_name": order.job_name,
            "customer": order.customer,
            "items": [
                {
                    "item": item.item,
                    "width_text": item.width_text,
                    "height_text": item.height_text,
                    "delivery_date": item.delivery_date,
                    "customer": item.customer,
                    "processing": list(item.processing),
                    "machine_hints": list(item.machine_hints),
                    "rows": list(item.rows),
                }
                for item in order.items.values()
            ],
        }
        for order in orders
    ]


def process_orders_from_cache(records: object) -> list[ProcessOrder]:
    if not isinstance(records, list):
        return []
    orders: list[ProcessOrder] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        aw_order = str(record.get("aw_order", "")).strip()
        job_name = str(record.get("job_name", "")).strip()
        if not aw_order or not job_name:
            continue
        order = ProcessOrder(aw_order, job_name, str(record.get("customer", "")))
        raw_items = record.get("items", [])
        if isinstance(raw_items, list):
            for raw_item in raw_items:
                if not isinstance(raw_item, dict):
                    continue
                try:
                    item_number = int(raw_item.get("item", 0))
                except (TypeError, ValueError):
                    continue
                if item_number <= 0:
                    continue
                raw_processing = raw_item.get("processing", [])
                raw_machine_hints = raw_item.get("machine_hints", [])
                raw_rows = raw_item.get("rows", [])
                if not isinstance(raw_processing, list):
                    raw_processing = []
                if not isinstance(raw_machine_hints, list):
                    raw_machine_hints = []
                if not isinstance(raw_rows, list):
                    raw_rows = []
                item = ProcessItem(
                    item=item_number,
                    width_text=str(raw_item.get("width_text", "")),
                    height_text=str(raw_item.get("height_text", "")),
                    delivery_date=str(raw_item.get("delivery_date", "")),
                    customer=str(raw_item.get("customer", "")),
                    processing=[str(value) for value in raw_processing if str(value)],
                    machine_hints=[str(value) for value in raw_machine_hints if str(value)],
                    rows=[int(value) for value in raw_rows if str(value).isdigit()],
                )
                order.items[item_number] = item
        orders.append(order)
    return orders


def load_process_orders_from_workbook(path: Path) -> list[ProcessOrder]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active

    try:
        return load_process_orders_from_rows(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def load_process_orders_from_legacy_xls(
    path: Path,
    progress_callback: ProcessListProgress | None = None,
) -> list[ProcessOrder]:
    raw_prefix = read_file_prefix(path, 512)
    if raw_prefix.lstrip().startswith(b"\xd0\xcf\x11\xe0"):
        converted = convert_legacy_xls_to_xlsx(path, progress_callback)
        return load_process_orders_from_workbook(converted)

    text_prefix = decode_text_file(path)[:2048].lstrip().lower()
    if text_prefix.startswith("<?xml") or text_prefix.startswith("<workbook"):
        return load_process_orders_from_rows(load_rows_from_spreadsheet_xml(path))
    if text_prefix.startswith("{\\rtf"):
        return load_process_orders_from_rows(load_rows_from_rtf(path))
    if text_prefix.startswith("<html") or "<table" in text_prefix:
        return load_process_orders_from_rows(load_rows_from_html_table(path))
    raise RuntimeError(
        f"{path.name} is not a supported .xls process-list export. "
        "Use Excel XML (.xml), RTF (.rtf), or the current .xlsx format."
    )


def converted_xlsx_source_hash_path(target: Path) -> Path:
    """Sidecar used to reuse a converted workbook after timestamp-only source changes."""
    return Path(target).with_name(f"{Path(target).name}.source.sha256")


def legacy_xls_conversion_cache_matches(path: Path, target: Path) -> bool:
    """Return True when an existing conversion still represents the XLS content.

    Network recopies commonly refresh timestamps even when the process list bytes
    are unchanged. Comparing a cached source hash avoids paying Excel startup and
    SaveAs costs again for the same logical workbook.
    """
    path = Path(path)
    target = Path(target)
    if not target.exists():
        return False
    try:
        if target.stat().st_mtime_ns >= path.stat().st_mtime_ns:
            return True
    except OSError:
        return False
    sidecar = converted_xlsx_source_hash_path(target)
    try:
        expected = sidecar.read_text(encoding="ascii").strip().casefold()
    except OSError:
        return False
    if not re.fullmatch(r"[0-9a-f]{64}", expected):
        return False
    try:
        current = shower_cache.cached_file_sha256("legacy_xls_conversion_source_v1", path).casefold()
    except OSError:
        return False
    return current == expected


def remember_legacy_xls_conversion_source(path: Path, target: Path) -> None:
    """Persist the exact XLS content hash represented by a converted XLSX."""
    try:
        digest = shower_cache.cached_file_sha256("legacy_xls_conversion_source_v1", Path(path))
        sidecar = converted_xlsx_source_hash_path(Path(target))
        sidecar.parent.mkdir(parents=True, exist_ok=True)
        temporary = sidecar.with_name(f".{sidecar.name}.{os.getpid()}.tmp")
        temporary.write_text(digest, encoding="ascii")
        os.replace(temporary, sidecar)
    except OSError:
        # Conversion remains valid even if the optimization sidecar cannot be written.
        return


def legacy_xls_excel_conversion_script(source_text: str, target_text: str) -> str:
    """Return a low-overhead Excel COM conversion script for binary XLS files."""
    return (
        "$Source = @'\n"
        + source_text
        + "\n'@\n$Target = @'\n"
        + target_text
        + "\n'@\n"
        + r"""
$ErrorActionPreference = 'Stop'
$excel = $null
$workbook = $null
try {
  $excel = New-Object -ComObject Excel.Application
  $excel.Visible = $false
  $excel.DisplayAlerts = $false
  $excel.ScreenUpdating = $false
  $excel.EnableEvents = $false
  $excel.AskToUpdateLinks = $false
  try { $excel.AutomationSecurity = 3 } catch {}
  try { $excel.Calculation = -4135 } catch {}
  # UpdateLinks=0 and ReadOnly=$true avoid unnecessary link refresh/write work
  # while the legacy process list is being normalized for read-only parsing.
  $workbook = $excel.Workbooks.Open($Source, 0, $true)
  try { $workbook.CheckCompatibility = $false } catch {}
  $workbook.SaveAs($Target, 51)
  $workbook.Close($false)
} finally {
  if ($workbook -ne $null) {
    [System.Runtime.Interopservices.Marshal]::ReleaseComObject($workbook) | Out-Null
  }
  if ($excel -ne $null) {
    $excel.Quit()
    [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
  }
}
"""
    )


def convert_legacy_xls_to_xlsx(
    path: Path,
    progress_callback: ProcessListProgress | None = None,
) -> Path:
    target = converted_xlsx_path(path)
    if legacy_xls_conversion_cache_matches(path, target):
        if progress_callback:
            progress_callback("normalized", path, f"Using cached {target.name}; Excel conversion skipped")
        return target

    target.parent.mkdir(parents=True, exist_ok=True)
    source_text = str(path.resolve())
    timeout_seconds = 75
    attempts = 2
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        staging = target.with_name(f".{target.stem}.attempt{attempt}.xlsx")
        try:
            staging.unlink(missing_ok=True)
        except OSError:
            pass
        target_text = str(staging.resolve())
        script = legacy_xls_excel_conversion_script(source_text, target_text)
        started = time.monotonic()
        if progress_callback:
            progress_callback(
                "normalizing",
                path,
                f"Opening legacy XLS in hidden Excel (attempt {attempt}/{attempts}; links/events disabled)",
            )
        excel_processes_before = excel_process_ids()
        try:
            subprocess.run(
                hidden_powershell_command(script, bypass_execution_policy=True),
                check=True,
                capture_output=True,
                text=True,
                timeout=timeout_seconds,
                **hidden_windows_subprocess_options(),
            )
            if not staging.exists():
                raise RuntimeError("Excel completed without creating the converted workbook")
            staging.replace(target)
            remember_legacy_xls_conversion_source(path, target)
            elapsed = time.monotonic() - started
            if progress_callback:
                progress_callback("normalized", path, f"Created {target.name} in {elapsed:.1f}s")
            return target
        except Exception as exc:
            last_error = exc
            stop_excel_processes(excel_process_ids() - excel_processes_before)
            try:
                staging.unlink(missing_ok=True)
            except OSError:
                pass
            if progress_callback:
                elapsed = time.monotonic() - started
                progress_callback("retry", path, f"Conversion attempt {attempt} did not finish after {elapsed:.1f}s")

    detail = (
        f"Excel did not finish auto-converting it after {attempts} attempts of {timeout_seconds} seconds"
        if isinstance(last_error, subprocess.TimeoutExpired)
        else "Excel could not auto-convert it"
    )
    raise RuntimeError(
        f"{path.name} is a binary Excel 97-2003 .xls file and {detail}. "
        "Use the Crystal Reports XML export or save the workbook as .xlsx."
    ) from last_error


def preconvert_legacy_xls_files(
    paths: Iterable[Path],
    progress_callback: ProcessListProgress | None = None,
) -> dict[Path, Path]:
    """Convert new local binary XLS exports in one hidden Excel session."""
    converted: dict[Path, Path] = {}
    pending: list[tuple[Path, Path, Path]] = []
    for raw_path in paths:
        path = Path(raw_path).resolve()
        if path.suffix.lower() != ".xls":
            continue
        try:
            if not read_file_prefix(path, 8).lstrip().startswith(b"\xd0\xcf\x11\xe0"):
                continue
        except OSError:
            continue
        cached_orders = shower_cache.load(PROCESS_LIST_CACHE_NAMESPACE, path)
        if (
            isinstance(cached_orders, dict)
            and cached_orders.get("schema") == PROCESS_LIST_CACHE_SCHEMA
            and cached_orders.get("orders")
        ):
            continue
        target = converted_xlsx_path(path)
        if legacy_xls_conversion_cache_matches(path, target):
            converted[path] = target
            if progress_callback:
                progress_callback("normalized", path, f"Using cached {target.name}; Excel conversion skipped")
            continue
        target.parent.mkdir(parents=True, exist_ok=True)
        staging = target.with_name(f".{target.stem}.batch.xlsx")
        pending.append((path, target, staging))

    if not pending:
        return converted
    if len(pending) == 1:
        path, _target, _staging = pending[0]
        converted[path] = convert_legacy_xls_to_xlsx(path, progress_callback)
        return converted

    def ps_literal(value: Path) -> str:
        return str(value.resolve()).replace("'", "''")

    jobs = ",\n".join(
        "@{ Source = '" + ps_literal(path) + "'; Target = '" + ps_literal(staging) + "' }"
        for path, _target, staging in pending
    )
    script = f"""
$ErrorActionPreference = 'Stop'
$jobs = @(
{jobs}
)
$excel = $null
try {{
  $excel = New-Object -ComObject Excel.Application
  $excel.Visible = $false
  $excel.DisplayAlerts = $false
  $excel.ScreenUpdating = $false
  $excel.EnableEvents = $false
  $excel.AskToUpdateLinks = $false
  try {{ $excel.AutomationSecurity = 3 }} catch {{}}
  try {{ $excel.Calculation = -4135 }} catch {{}}
  foreach ($job in $jobs) {{
    $workbook = $null
    try {{
      $workbook = $excel.Workbooks.Open($job.Source, 0, $true)
      try {{ $workbook.CheckCompatibility = $false }} catch {{}}
      $workbook.SaveAs($job.Target, 51)
      $workbook.Close($false)
    }} finally {{
      if ($workbook -ne $null) {{
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($workbook) | Out-Null
      }}
    }}
  }}
}} finally {{
  if ($excel -ne $null) {{
    $excel.Quit()
    [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
  }}
}}
"""
    for path, _target, staging in pending:
        try:
            staging.unlink(missing_ok=True)
        except OSError:
            pass
        if progress_callback:
            progress_callback("normalizing", path, f"Queued local XLS conversion ({len(pending)} files, one Excel session)")

    excel_processes_before = excel_process_ids()
    try:
        subprocess.run(
            hidden_powershell_command(script, bypass_execution_policy=True),
            check=True,
            capture_output=True,
            text=True,
            timeout=max(90, 35 * len(pending)),
            **hidden_windows_subprocess_options(),
        )
        for path, target, staging in pending:
            if not staging.exists():
                raise RuntimeError(f"Excel did not create {staging.name}")
            staging.replace(target)
            remember_legacy_xls_conversion_source(path, target)
            converted[path] = target
            if progress_callback:
                progress_callback("normalized", path, f"Created {target.name} locally")
        return converted
    except Exception:
        stop_excel_processes(excel_process_ids() - excel_processes_before)
        for _path, _target, staging in pending:
            try:
                staging.unlink(missing_ok=True)
            except OSError:
                pass
        # Preserve the established per-file retry and error detail if the fast
        # shared Excel session cannot normalize the complete batch.
        for path, _target, _staging in pending:
            converted[path] = convert_legacy_xls_to_xlsx(path, progress_callback)
        return converted


def converted_xlsx_path(path: Path) -> Path:
    digest = hashlib.sha1(str(path.resolve()).encode("utf-8", errors="ignore")).hexdigest()[:10]
    cache_root = shower_cache.configured_root()
    output_root = cache_root.parent if cache_root is not None else programmer.default_output_dir()
    return output_root / "Converted Process Lists" / f"{path.stem}_{digest}.xlsx"


def hidden_powershell_command(script: str, *, bypass_execution_policy: bool = False) -> list[str]:
    command = [
        "powershell",
        "-NoProfile",
        "-NonInteractive",
        "-WindowStyle",
        "Hidden",
    ]
    if bypass_execution_policy:
        command.extend(["-ExecutionPolicy", "Bypass"])
    command.extend(["-Command", script])
    return command


def hidden_windows_subprocess_options() -> dict[str, object]:
    if os.name != "nt":
        return {}
    options: dict[str, object] = {
        "creationflags": getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000),
    }
    startupinfo_type = getattr(subprocess, "STARTUPINFO", None)
    if startupinfo_type is not None:
        startupinfo = startupinfo_type()
        startupinfo.dwFlags |= getattr(subprocess, "STARTF_USESHOWWINDOW", 0x00000001)
        startupinfo.wShowWindow = 0
        options["startupinfo"] = startupinfo
    return options


def excel_process_ids() -> set[int]:
    try:
        result = subprocess.run(
            hidden_powershell_command(
                "Get-Process EXCEL -ErrorAction SilentlyContinue | ForEach-Object { $_.Id }"
            ),
            check=True,
            capture_output=True,
            text=True,
            timeout=5,
            **hidden_windows_subprocess_options(),
        )
    except Exception:
        return set()
    ids: set[int] = set()
    for line in result.stdout.splitlines():
        line = line.strip()
        if line.isdigit():
            ids.add(int(line))
    return ids


def stop_excel_processes(process_ids: set[int]) -> None:
    if not process_ids:
        return
    id_text = ",".join(str(process_id) for process_id in sorted(process_ids))
    try:
        subprocess.run(
            hidden_powershell_command(
                f"Stop-Process -Id {id_text} -Force -ErrorAction SilentlyContinue"
            ),
            check=False,
            capture_output=True,
            text=True,
            timeout=10,
            **hidden_windows_subprocess_options(),
        )
    except Exception:
        pass


def process_list_is_mirror_batch(rows: Iterable[Iterable[object]]) -> bool:
    """Identify a mirror-only material list without matching customer or job names."""
    material_descriptions: list[str] = []
    for row in rows:
        for value in row:
            text = cell_text(value)
            if GLASS_MATERIAL_PATTERN.fullmatch(text):
                material_descriptions.append(text.upper())
    return bool(material_descriptions) and all("MIRROR" in text for text in material_descriptions)


def mirror_waterjet_orders(orders: Iterable[ProcessOrder]) -> list[ProcessOrder]:
    """Keep only items routed through the Waterjet section of a mirror batch."""
    waterjet_orders: list[ProcessOrder] = []
    for order in orders:
        waterjet_items = {
            item_number: item
            for item_number, item in order.items.items()
            if item.desired_machine() == "WJ"
        }
        if not waterjet_items:
            continue
        order.items = waterjet_items
        waterjet_orders.append(order)
    return waterjet_orders


def load_process_orders_from_rows(rows: Iterable[Iterable[object]]) -> list[ProcessOrder]:
    materialized_rows = [list(row) for row in rows]
    mirror_batch = process_list_is_mirror_batch(materialized_rows)
    orders: dict[tuple[str, str], ProcessOrder] = {}
    last_key: tuple[str, str, int] | None = None
    for row_number, values in enumerate(materialized_rows, start=1):
        order_item = cell_at(values, 6)
        job_name = programmer.clean_job_name(cell_at(values, 13))
        customer = cell_at(values, 10)
        width_text = cell_at(values, 2)
        height_text = cell_at(values, 3)
        processing = cell_at(values, 7)
        delivery_date = cell_at(values, 8)
        machine_hint = cell_at(values, 21)

        parsed = parse_order_item(order_item)
        if parsed:
            aw_order, item_number = parsed
            key = (aw_order, job_name)
            order = orders.get(key)
            if order is None:
                order = ProcessOrder(aw_order=aw_order, job_name=job_name, customer=customer)
                orders[key] = order
            if customer and not order.customer:
                order.customer = customer
            item = order.items.get(item_number)
            if item is None:
                item = ProcessItem(item=item_number)
                order.items[item_number] = item
            item.add_row(row_number, width_text, height_text, delivery_date, customer, processing, machine_hint)
            last_key = (aw_order, job_name, item_number)
            continue

        if last_key and processing and job_name:
            aw_order, previous_job, item_number = last_key
            if job_name != previous_job:
                continue
            order = orders[(aw_order, previous_job)]
            order.items[item_number].add_row(row_number, width_text, height_text, delivery_date, customer, processing, machine_hint)

    loaded_orders = sorted(orders.values(), key=lambda order: int(order.aw_order))
    return mirror_waterjet_orders(loaded_orders) if mirror_batch else loaded_orders


def load_rows_from_spreadsheet_xml(path: Path) -> list[list[str]]:
    root = ET.parse(path).getroot()
    for table in root.iter():
        if xml_local_name(table.tag) != "Table":
            continue
        rows = parse_spreadsheet_xml_table(table)
        if rows:
            return rows
    return []


def load_rows_from_crystal_xml(path: Path) -> list[list[str]]:
    root = ET.parse(path).getroot()
    if "crystal-reports" not in root.tag.lower():
        return []

    records: list[dict[str, object]] = []
    current: dict[str, object] | None = None
    pending: dict[str, object] = {}

    def finish_current() -> None:
        nonlocal current
        if current and current.get("order_item"):
            records.append(current)
        current = None

    for field in root.iter():
        if xml_local_name(field.tag) != "Field":
            continue
        field_name = field.attrib.get("FieldName", "")
        value = crystal_field_value(field)
        if not value:
            continue

        if field_name == "{@Processingtext}":
            if current and current.get("order_item") and current.get("width") and current.get("height"):
                finish_current()
            pending = {"processing": value}
            continue
        if field_name == "{PROD_JOBITEM_GLASS.MENGE}":
            pending["quantity"] = value
            continue
        if field_name == "{@order_item}":
            finish_current()
            current = dict(pending)
            current["order_item"] = value
            pending = {}
            continue
        if current is None:
            continue

        if field_name == "{@Sheet_Width}":
            current["width"] = value
        elif field_name == "{@Sheet_Height}":
            current["height"] = value
        elif field_name == "{BW_AUFTR_KOPF.BEST_TEXT2}":
            current["delivery_date"] = value
        elif field_name == "{BW_AUFTR_KOPF.AH_NAME1}":
            current["customer"] = value
        elif field_name == "{BW_AUFTR_KOPF.BEST_TEXT1}":
            current["job_name"] = value
        elif field_name == "{@Next_machine}":
            machine_hints = current.setdefault("machine_hints", [])
            if isinstance(machine_hints, list) and value not in machine_hints:
                machine_hints.append(value)

    finish_current()
    return [crystal_record_to_process_row(record) for record in records]


def crystal_field_value(field: ET.Element) -> str:
    for child in field:
        if xml_local_name(child.tag) == "FormattedValue":
            return cell_text("".join(child.itertext()))
    for child in field:
        if xml_local_name(child.tag) == "Value":
            return cell_text("".join(child.itertext()))
    return cell_text("".join(field.itertext()))


def crystal_record_to_process_row(record: dict[str, object]) -> list[str]:
    row = [""] * 22
    row[2] = str(record.get("width", ""))
    row[3] = str(record.get("height", ""))
    row[6] = str(record.get("order_item", ""))
    row[7] = str(record.get("processing", ""))
    row[8] = str(record.get("delivery_date", ""))
    row[10] = str(record.get("customer", ""))
    row[13] = str(record.get("job_name", ""))
    machine_hints = record.get("machine_hints", [])
    if isinstance(machine_hints, list):
        row[21] = " | ".join(str(value) for value in machine_hints)
    return row


def parse_spreadsheet_xml_table(table: ET.Element) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in table:
        if xml_local_name(row.tag) != "Row":
            continue
        values: list[str] = []
        current_col = 0
        for cell in row:
            if xml_local_name(cell.tag) != "Cell":
                continue
            index_text = xml_attr(cell, "Index")
            if index_text and index_text.isdigit():
                current_col = max(int(index_text) - 1, 0)
            while len(values) < current_col:
                values.append("")
            values.append(xml_cell_text(cell))
            current_col += 1
        if any(value for value in values):
            rows.append(values)
    return rows


def xml_cell_text(cell: ET.Element) -> str:
    for child in cell:
        if xml_local_name(child.tag) == "Data":
            return cell_text("".join(child.itertext()))
    return cell_text("".join(cell.itertext()))


def xml_attr(element: ET.Element, local_name: str) -> str:
    for key, value in element.attrib.items():
        if xml_local_name(key) == local_name:
            return value
    return ""


def xml_local_name(name: str) -> str:
    return name.rsplit("}", 1)[-1]


def load_rows_from_rtf(path: Path) -> list[list[str]]:
    text = rtf_to_text(decode_text_file(path))
    rows: list[list[str]] = []
    for line in text.splitlines():
        values = [cell_text(part) for part in line.split("\t")]
        if any(values):
            rows.append(values)
    return rows


def load_rows_from_crystal_rtf(path: Path) -> list[list[str]]:
    lines = [
        cell_text(line)
        for line in rtf_to_text(decode_text_file(path)).splitlines()
        if cell_text(line)
    ]
    rows: list[list[str]] = []
    for index, line in enumerate(lines):
        parsed = parse_order_item(line)
        if not parsed:
            continue
        measurements = measurement_lines_after(lines, index + 1, 2)
        width = measurements[0] if len(measurements) >= 1 else ""
        height = measurements[1] if len(measurements) >= 2 else ""
        machine_hint = first_matching_line(lines[index + 1:index + 8], r"\b(?:WATER\s*JET|WATERJET|WJ|DENVER\s*[12])\b")
        date_index = first_matching_index(lines, index + 1, index + 10, r"\b\d{1,2}/\d{1,2}/\d{2,4}\b")
        delivery_date = lines[date_index] if date_index is not None else ""
        customer = lines[date_index + 1] if date_index is not None and date_index + 1 < len(lines) else ""
        job_name = lines[date_index + 2] if date_index is not None and date_index + 2 < len(lines) else ""
        processing = " | ".join(
            value
            for value in lines[max(0, index - 4):index]
            if not re.fullmatch(r"\d+(?:\.\d+)?", value)
        )
        row = [""] * 22
        row[2] = width
        row[3] = height
        row[6] = line.replace("_", "-")
        row[7] = processing
        row[8] = delivery_date
        row[10] = customer
        row[13] = programmer.clean_job_name(job_name)
        row[21] = machine_hint
        rows.append(row)
    return rows


def measurement_lines_after(lines: list[str], start: int, count: int) -> list[str]:
    measurements: list[str] = []
    for line in lines[start:start + 8]:
        if re.search(r'\d+\s*"\s*(?:\d+/\d+)?|\d+-\d+/\d+|\d+/\d+', line):
            measurements.append(line)
            if len(measurements) >= count:
                break
    return measurements


def first_matching_line(lines: Iterable[str], pattern: str) -> str:
    regex = re.compile(pattern, re.IGNORECASE)
    for line in lines:
        if regex.search(line):
            return line
    return ""


def first_matching_index(lines: list[str], start: int, stop: int, pattern: str) -> int | None:
    regex = re.compile(pattern)
    for index in range(start, min(stop, len(lines))):
        if regex.search(lines[index]):
            return index
    return None


def rtf_to_text(source: str) -> str:
    result: list[str] = []
    index = 0
    while index < len(source):
        char = source[index]
        if char in "{}":
            index += 1
            continue
        if char != "\\":
            result.append(char)
            index += 1
            continue

        index += 1
        if index >= len(source):
            break
        escaped = source[index]
        if escaped in "{}\\":
            result.append(escaped)
            index += 1
            continue
        if escaped == "'":
            hex_value = source[index + 1:index + 3]
            if len(hex_value) == 2:
                try:
                    result.append(bytes.fromhex(hex_value).decode("cp1252"))
                except Exception:
                    pass
            index += 3
            continue
        if escaped in "\r\n":
            index += 1
            continue

        start = index
        while index < len(source) and source[index].isalpha():
            index += 1
        word = source[start:index]
        sign = 1
        if index < len(source) and source[index] == "-":
            sign = -1
            index += 1
        number_start = index
        while index < len(source) and source[index].isdigit():
            index += 1
        number = source[number_start:index]
        if index < len(source) and source[index] == " ":
            index += 1

        if word == "cell":
            result.append("\t")
        elif word == "row":
            result.append("\n")
        elif word in {"par", "line"}:
            result.append("\n")
        elif word == "tab":
            result.append("\t")
        elif word == "u" and number:
            codepoint = sign * int(number)
            if codepoint < 0:
                codepoint += 65536
            try:
                result.append(chr(codepoint))
            except ValueError:
                pass
    return "".join(result)


class ProcessListHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self.current_row: list[str] | None = None
        self.current_cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "tr":
            self.current_row = []
        elif tag.lower() in {"td", "th"} and self.current_row is not None:
            self.current_cell = []

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"td", "th"} and self.current_row is not None and self.current_cell is not None:
            self.current_row.append(cell_text(html.unescape("".join(self.current_cell))))
            self.current_cell = None
        elif tag.lower() == "tr" and self.current_row is not None:
            if any(self.current_row):
                self.rows.append(self.current_row)
            self.current_row = None
            self.current_cell = None

    def handle_data(self, data: str) -> None:
        if self.current_cell is not None:
            self.current_cell.append(data)


def load_rows_from_html_table(path: Path) -> list[list[str]]:
    parser = ProcessListHTMLParser()
    parser.feed(decode_text_file(path))
    return parser.rows


def merge_process_order(
    target_orders: dict[tuple[str, str], ProcessOrder],
    order: ProcessOrder,
) -> None:
    key = (order.aw_order, order.job_name)
    target = target_orders.get(key)
    if target is None:
        target_orders[key] = order
        return
    if order.customer and not target.customer:
        target.customer = order.customer
    for item_number, item in order.items.items():
        target_item = target.items.get(item_number)
        if target_item is None:
            target.items[item_number] = item
            continue
        if item.width_text and not target_item.width_text:
            target_item.width_text = item.width_text
        if item.height_text and not target_item.height_text:
            target_item.height_text = item.height_text
        if item.delivery_date and not target_item.delivery_date:
            target_item.delivery_date = item.delivery_date
        if item.customer and not target_item.customer:
            target_item.customer = item.customer
        for processing in item.processing:
            if processing not in target_item.processing:
                target_item.processing.append(processing)
        for machine_hint in item.machine_hints:
            if machine_hint not in target_item.machine_hints:
                target_item.machine_hints.append(machine_hint)
        target_item.rows.extend(item.rows)




def merge_process_orders_by_aw(orders: Iterable[ProcessOrder]) -> list[ProcessOrder]:
    """Combine split process-list records using A&W order as the durable identity."""
    merged: dict[str, ProcessOrder] = {}
    sequence: list[str] = []
    for order in orders:
        aw_order = str(order.aw_order).strip()
        if not aw_order:
            continue
        target = merged.get(aw_order)
        if target is None:
            target = clone_process_order(order)
            merged[aw_order] = target
            sequence.append(aw_order)
            continue
        if not target.job_name and order.job_name:
            target.job_name = order.job_name
        if not target.customer and order.customer:
            target.customer = order.customer
        for item_number, item in order.items.items():
            target_item = target.items.get(item_number)
            if target_item is None:
                target.items[item_number] = ProcessItem(
                    item=item.item,
                    width_text=item.width_text,
                    height_text=item.height_text,
                    delivery_date=item.delivery_date,
                    customer=item.customer,
                    processing=list(item.processing),
                    machine_hints=list(item.machine_hints),
                    rows=list(item.rows),
                )
                continue
            for field_name in ("width_text", "height_text", "delivery_date", "customer"):
                if not getattr(target_item, field_name) and getattr(item, field_name):
                    setattr(target_item, field_name, getattr(item, field_name))
            for value in item.processing:
                if value not in target_item.processing:
                    target_item.processing.append(value)
            for value in item.machine_hints:
                if value not in target_item.machine_hints:
                    target_item.machine_hints.append(value)
            target_item.rows.extend(item.rows)
    return [merged[aw_order] for aw_order in sequence]


def unique_orders_from_batches(batches: list[dict[str, object]]) -> list[ProcessOrder]:
    """Return one merged ProcessOrder per A&W order across visible batches."""
    ordered: list[ProcessOrder] = []
    for batch in batches:
        batch_orders = batch.get("orders", [])
        if isinstance(batch_orders, list):
            ordered.extend(order for order in batch_orders if isinstance(order, ProcessOrder))
    return merge_process_orders_by_aw(ordered)

def clone_process_order(order: ProcessOrder) -> ProcessOrder:
    cloned = ProcessOrder(aw_order=order.aw_order, job_name=order.job_name, customer=order.customer)
    for item_number, item in order.items.items():
        cloned.items[item_number] = ProcessItem(
            item=item.item,
            width_text=item.width_text,
            height_text=item.height_text,
            delivery_date=item.delivery_date,
            customer=item.customer,
            processing=list(item.processing),
            machine_hints=list(item.machine_hints),
            rows=list(item.rows),
        )
    return cloned


def cell_at(values: list[object], index: int) -> str:
    if index >= len(values):
        return ""
    return cell_text(values[index])


def filter_orders(orders: Iterable[ProcessOrder], requested: str | None) -> list[ProcessOrder]:
    if not requested:
        return list(orders)
    wanted = {part.strip() for part in requested.split(",") if part.strip()}
    wanted_jobs = {programmer.normalize_lookup(part) for part in wanted}
    return [order for order in orders if order.aw_order in wanted or programmer.normalize_lookup(order.job_name) in wanted_jobs]


def visible_orders(orders: Iterable[ProcessOrder], config: dict[str, object]) -> list[ProcessOrder]:
    visible: list[ProcessOrder] = []
    for order in orders:
        if order.is_mirror(config) and not order.has_mirror_fabrication(config):
            continue
        visible.append(order)
    return visible


def attach_transom_panels(
    reader: PdfReader,
    panels: list[programmer.Panel],
    process_order: ProcessOrder,
    config: dict[str, object],
) -> None:
    existing_items = {panel.item for panel in panels}
    missing_items = [item for item in process_order.item_numbers if item not in existing_items]
    if not missing_items:
        return

    used_pages = {panel.page_index for panel in panels}
    candidates: list[tuple[int, str, str]] = []
    for page_index, page in enumerate(reader.pages):
        if page_index == 0 or page_index in used_pages:
            continue
        text = page.extract_text() or ""
        if not text.strip() or programmer.looks_like_template_page(text):
            continue
        if programmer.extract_item_number(text) is not None:
            continue
        transom_label = programmer.extract_transom_label(text)
        if transom_label is None:
            continue
        candidates.append((page_index, text, transom_label))

    if not candidates:
        return

    candidates.sort(key=lambda entry: entry[0])
    missing_items = sorted(missing_items)
    pair_count = min(len(missing_items), len(candidates))
    for item_number, (page_index, text, transom_label) in zip(missing_items[-pair_count:], candidates[-pair_count:]):
        width, height = programmer.extract_dimensions(text)
        panel = programmer.classify_panel(
            programmer.Panel(
                item=item_number,
                page_index=page_index,
                text=text,
                width=width,
                height=height,
                machine="",
            ),
            config,
            process_order.aw_order,
        )
        panel.reasons.append(f"transom sketch label {transom_label} mapped to P{item_number}")
        panels.append(panel)

    panels.sort(key=lambda panel: panel.item)


def attach_unlabeled_process_pages(
    reader: PdfReader,
    panels: list[programmer.Panel],
    process_order: ProcessOrder,
    config: dict[str, object],
) -> None:
    existing_items = {panel.item for panel in panels}
    missing_items = [item for item in process_order.item_numbers if item not in existing_items]
    if not missing_items:
        return

    used_pages = {panel.page_index for panel in panels}
    candidates: list[tuple[int, str]] = []
    for page_index, page in enumerate(reader.pages):
        if page_index == 0 or page_index in used_pages:
            continue
        text = page.extract_text() or ""
        if not text.strip() or programmer.looks_like_template_page(text):
            continue
        if programmer.extract_item_number(text) is not None:
            continue
        width, height = programmer.extract_dimensions(text)
        if width is None or height is None:
            continue
        candidates.append((page_index, text))

    for item_number, (page_index, text) in zip(sorted(missing_items), candidates):
        width, height = programmer.extract_dimensions(text)
        panel = programmer.classify_panel(
            programmer.Panel(
                item=item_number,
                page_index=page_index,
                text=text,
                width=width,
                height=height,
                machine="",
            ),
            config,
            process_order.aw_order,
        )
        panel.reasons.append(f"unlabeled piece page mapped to P{item_number} from process list")
        panels.append(panel)

    panels.sort(key=lambda panel: panel.item)


def attach_unmarked_process_pages(
    reader: PdfReader,
    panels: list[programmer.Panel],
    process_order: ProcessOrder,
    config: dict[str, object],
) -> None:
    """Map piece pages that contain no readable P-number, dimensions, or text.

    The process list is the authority for item order. After all stronger matching
    strategies run, remaining non-template pages are paired with remaining items
    in page order. This lets a completely unmarked sketch remain valid while the
    DXF is still programmed from process-list data.
    """
    existing_items = {panel.item for panel in panels}
    missing_items = [item for item in process_order.item_numbers if item not in existing_items]
    if not missing_items:
        return

    used_pages = {panel.page_index for panel in panels}
    candidates: list[tuple[int, str]] = []
    for page_index, page in enumerate(reader.pages):
        if page_index == 0 or page_index in used_pages:
            continue
        text_value = page.extract_text() or ""
        if programmer.looks_like_template_page(text_value):
            continue
        if programmer.extract_item_number(text_value) is not None:
            continue
        candidates.append((page_index, text_value))

    # Some vendors export a piece-only PDF without a cover page. Use page 0 as
    # a final fallback only when the file does not contain enough later pages
    # to represent the process-list items.
    if len(candidates) < len(missing_items) and len(reader.pages) <= len(missing_items):
        page_index = 0
        if page_index not in used_pages and reader.pages:
            page_zero_text = reader.pages[0].extract_text() or ""
            if (
                not programmer.looks_like_template_page(page_zero_text)
                and programmer.extract_item_number(page_zero_text) is None
            ):
                candidates.insert(0, (page_index, page_zero_text))

    for item_number, (page_index, text_value) in zip(sorted(missing_items), candidates):
        width, height = programmer.extract_dimensions(text_value)
        panel = programmer.classify_panel(
            programmer.Panel(
                item=item_number,
                page_index=page_index,
                text=text_value,
                width=width,
                height=height,
                machine="",
            ),
            config,
            process_order.aw_order,
        )
        panel.reasons.append(f"unmarked sketch page mapped to P{item_number} from process list")
        panels.append(panel)

    panels.sort(key=lambda panel: panel.item)


def reconcile_process_list_item_gaps(
    panels: list[programmer.Panel],
    process_order: ProcessOrder,
) -> None:
    expected_items = sorted(process_order.item_numbers)
    ordered_panels = sorted(panels, key=lambda panel: (panel.page_index, panel.item))
    actual_items = [panel.item for panel in ordered_panels]
    if not expected_items or len(expected_items) != len(actual_items):
        return
    if set(expected_items) == set(actual_items):
        return
    missing = set(expected_items) - set(actual_items)
    extra = set(actual_items) - set(expected_items)
    if not missing or len(missing) != len(extra):
        return

    remaps: list[tuple[programmer.Panel, int, int]] = []
    for panel, expected_item in zip(ordered_panels, expected_items):
        if panel.item == expected_item:
            continue
        if expected_item not in missing or panel.item not in extra:
            return
        remaps.append((panel, panel.item, expected_item))
    if len(remaps) != len(missing):
        return

    for panel, original_item, process_item in remaps:
        panel.source_item = original_item
        panel.item = process_item
        panel.reasons.append(f"sketch P{original_item} mapped to process-list P{process_item}")
    panels.sort(key=lambda panel: panel.item)


def remap_process_items_to_sketch_pages(
    panels: list[programmer.Panel],
    process_order: ProcessOrder,
    remake_items: set[int] | None = None,
) -> dict[int, int]:
    if not panels or not process_order.items:
        return {}
    actual_items = sorted({panel.item for panel in panels})
    process_items = sorted(process_order.item_numbers)
    extra_sketch_items = [item for item in actual_items if item not in set(process_items)]
    if not extra_sketch_items:
        return {}

    remaps: dict[int, int] = {}
    used_targets: set[int] = set()
    missing_process_items = [item for item in process_items if item not in set(actual_items)]
    for missing_item in missing_process_items:
        target = choose_sketch_item_for_process_gap(missing_item, extra_sketch_items, used_targets)
        if target is None:
            continue
        remaps[missing_item] = target
        used_targets.add(target)

    for process_item, sketch_item in remaps.items():
        if process_item not in process_order.items or sketch_item in process_order.items:
            continue
        item = process_order.items.pop(process_item)
        item.item = sketch_item
        process_order.items[sketch_item] = item
        for panel in panels:
            if panel.item == sketch_item:
                panel.source_item = process_item
                reason = f"process-list P{process_item} applied to sketch P{sketch_item}"
                if reason not in panel.reasons:
                    panel.reasons.append(reason)
    return remaps


def match_process_items_to_sketch_pages(
    reader: PdfReader,
    panels: list[programmer.Panel],
    process_order: ProcessOrder,
    config: dict[str, object],
    remake_items: set[int] | None = None,
) -> dict[int, int]:
    # A transom's sketch label identifies the panel below it (TRN2), not its
    # A+W item number. Attach explicit transom pages to the highest unmatched
    # process-list items before generic P-number gap reconciliation can claim
    # those rows for an ordinary panel.
    attach_transom_panels(reader, panels, process_order, config)
    return remap_process_items_to_sketch_pages(panels, process_order, remake_items)


def choose_sketch_item_for_process_gap(
    process_item: int,
    extra_sketch_items: list[int],
    used_targets: set[int],
) -> int | None:
    candidates = [item for item in extra_sketch_items if item not in used_targets]
    if not candidates:
        return None
    higher = [item for item in candidates if item > process_item]
    if higher:
        return min(higher, key=lambda item: (item - process_item, item))
    return min(candidates, key=lambda item: (abs(item - process_item), item))


def reconcile_missing_items_from_extra_sketch_pages(
    panels: list[programmer.Panel],
    process_order: ProcessOrder,
) -> None:
    expected_items = sorted(process_order.item_numbers)
    if not expected_items:
        return
    actual_items = {panel.item for panel in panels}
    missing_items = sorted(set(expected_items) - actual_items)
    if not missing_items:
        return
    extra_panels = sorted(
        (panel for panel in panels if panel.item not in set(expected_items)),
        key=lambda panel: (panel.page_index, panel.item),
    )
    if not extra_panels:
        return

    for missing_item in missing_items:
        higher = [panel for panel in extra_panels if panel.item > missing_item]
        if higher:
            chosen = min(higher, key=lambda panel: (panel.item - missing_item, panel.page_index))
        else:
            chosen = min(extra_panels, key=lambda panel: (abs(panel.item - missing_item), panel.page_index))
        original_item = chosen.item
        chosen.source_item = original_item
        chosen.item = missing_item
        chosen.reasons.append(
            f"sketch P{original_item} mapped to process-list P{missing_item}; missing process-list item matched to extra sketch page"
        )
        extra_panels.remove(chosen)
        if not extra_panels:
            break
    panels.sort(key=lambda panel: panel.item)


def apply_process_hints(
    panels: list[programmer.Panel],
    process_order: ProcessOrder,
    config: dict[str, object],
) -> None:
    for panel in panels:
        process_item = process_order.items.get(panel.item)
        if process_item is None:
            panel.warnings.append("No matching process-list item row found.")
            continue
        panel.process_text = process_item.text_blob()
        apply_process_dimensions(panel, process_item)
        dimension_notes = getattr(process_order, "dimension_match_notes", {})
        if isinstance(dimension_notes, dict):
            dimension_note = str(dimension_notes.get(panel.item, "")).strip()
            if dimension_note and dimension_note not in panel.reasons:
                panel.reasons.append(dimension_note)
        if bool(getattr(process_order, "manual_dimension_match_override_used", False)):
            warning = "Manual dimension-match override: operator accepted the process-list/sketch mismatch."
            if warning not in panel.warnings:
                panel.warnings.append(warning)

        process_glass_text = "\n".join((process_item.processing_text, process_item.machine_text))
        is_mirror_glass = (
            programmer.has_mirror_glass_type(panel.text, config)
            or programmer.has_mirror_glass_type(process_glass_text, config)
        )
        if is_mirror_glass:
            set_panel_machine(panel, "WJ", "mirror glass type always uses WJ")
            panel.indicator_corner = programmer.default_waterjet_indicator_corner(panel)
            panel.rotation_degrees = -90 if panel.height and panel.width and panel.height > panel.width else 0

        desired_machine = process_item.desired_machine()
        processing_machine = process_item.inferred_denver_machine(config)
        strong_process_wj = process_item.has_strong_waterjet_fabrication(config)
        strong_pdf_wj = programmer.has_pdf_waterjet_evidence(panel.text, config)
        has_door_evidence = programmer.has_door_programming_evidence(panel, config)
        if not is_mirror_glass and desired_machine:
            original_machine = panel.machine
            if desired_machine == "WJ":
                if strong_process_wj or strong_pdf_wj or denver_minimum_forces_wj(panel, config):
                    set_panel_machine(panel, desired_machine, f"process list machine: {desired_machine}")
                elif processing_machine:
                    panel.reasons.append("process list says WJ, but no WJ-only fabrication found")
                    set_panel_machine(panel, processing_machine, f"process-list fabrication suggests {processing_machine}")
                elif original_machine == "WJ":
                    panel.reasons.append(f"process list machine: {desired_machine}")
                else:
                    set_panel_machine(panel, desired_machine, f"process list machine: {desired_machine}")
            elif (strong_process_wj or strong_pdf_wj) and not has_door_evidence:
                set_panel_machine(
                    panel,
                    "WJ",
                    "WJ-only radius/notch fabrication overrides process-list Denver routing",
                )
            else:
                set_panel_machine(panel, desired_machine, f"process list machine: {desired_machine}")
        elif not is_mirror_glass and processing_machine:
            if panel.machine != processing_machine and (not panel.machine or panel.label_only):
                set_panel_machine(panel, processing_machine, f"process-list fabrication suggests {processing_machine}")
            else:
                panel.reasons.append(f"process-list fabrication confirms {processing_machine}")
        elif not is_mirror_glass:
            panel.machine = ""
            panel.label_only = True
            panel.skip_dxf = True
            panel.indicator_corner = None
            panel.rotation_degrees = None
            panel.reasons.append("process list has no cutting machine")

        if panel.machine == "WJ":
            panel.indicator_corner = programmer.default_waterjet_indicator_corner(panel)
            panel.rotation_degrees = -90 if panel.height and panel.width and panel.height > panel.width else 0
        elif panel.machine.startswith("DENVER"):
            panel.rotation_degrees = 90 if panel.height and panel.width and panel.height > panel.width else 0
            panel.indicator_corner = programmer.denver_grabber_corner_for_panel(panel, panel.rotation_degrees)

        if process_item.has_diamon_fusion:
            panel.diamon_fusion = True
            panel.reasons.append("process list Diamon Fusion")

        programmer.apply_auto_angle_correction(panel, config)
        programmer.apply_override(panel, config, process_order.aw_order)
        programmer.apply_auto_angle_correction(panel, config)
        programmer.validate_panel_constraints(panel, config)


def set_panel_machine(panel: programmer.Panel, machine: str, reason: str) -> None:
    panel.machine = machine
    panel.label_only = False
    panel.skip_dxf = False
    if reason not in panel.reasons:
        panel.reasons.append(reason)


def apply_process_dimensions(panel: programmer.Panel, process_item: ProcessItem) -> None:
    width = programmer.parse_measurement(process_item.width_text)
    height = programmer.parse_measurement(process_item.height_text)
    if width is None or height is None:
        return
    if panel.width != width or panel.height != height:
        panel.width = width
        panel.height = height
        panel.reasons.append("dimensions from process list")
    panel.warnings = [
        warning for warning in panel.warnings
        if warning != "Could not read dimensions from PDF text."
    ]


def apply_process_list_scope(panels: list[programmer.Panel], process_order: ProcessOrder) -> None:
    expected = set(process_order.item_numbers)
    for panel in panels:
        if panel.item in expected:
            continue
        panel.remake_excluded = True
        panel.machine = ""
        panel.label_only = True
        panel.skip_dxf = True
        panel.indicator_corner = None
        panel.rotation_degrees = None
        panel.angle_correction_degrees = 0.0
        panel.angle_correction_reason = ""
        panel.diamon_fusion = False
        panel.source_dxf = None
        panel.output_dxf = None
        panel.warnings = [
            warning
            for warning in panel.warnings
            if warning != "No matching process-list item row found."
        ]
        if "not in process list; X out" not in panel.reasons:
            panel.reasons.append("not in process list; X out")


def denver_minimum_forces_wj(panel: programmer.Panel, config: dict[str, object]) -> bool:
    rules = config.get("rules", {})
    if not isinstance(rules, dict):
        return False
    denver_min = float(rules.get("denver_min_inches", 6.125))
    return (
        panel.width is not None
        and panel.height is not None
        and min(panel.width, panel.height) < denver_min
    )


def process_list_fabrication_keywords(config: dict[str, object]) -> set[str]:
    keywords = upper_config_list(
        config,
        "fabrication_keywords",
        ["PPH", "GEN", "HINGE", "HOLE", "CUTOUT", "CUT-OUT", "CORNER NOTCH", "EDGE NOTCH", "NOTCH", "RADIUS"],
    )
    keywords.update(upper_config_list(config, "denver_fabrication_keywords", ["K CUT", "K-CUT"]))
    keywords.update(
        {
            "SLOT",
            "SLOTTED",
            "SCU",
            "SCU4",
            "MACRO",
            "CUT-IN",
            "CUT IN",
            "CUTIN",
        }
    )
    return keywords


def process_order_dimensions(process_order: ProcessOrder) -> list[tuple[int, float, float]]:
    dimensions: list[tuple[int, float, float]] = []
    for item_number, item in sorted(process_order.items.items()):
        width = programmer.parse_measurement(item.width_text)
        height = programmer.parse_measurement(item.height_text)
        if width is None or height is None:
            return []
        dimensions.append((item_number, width, height))
    return dimensions


def pdf_piece_dimensions(reader: PdfReader) -> list[tuple[int, float, float]]:
    dimensions: list[tuple[int, float, float]] = []
    for page_index, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if programmer.looks_like_template_page(text):
            continue
        width, height = programmer.extract_dimensions(text)
        if width is not None and height is not None:
            dimensions.append((page_index, width, height))
    return dimensions


def cached_pdf_piece_dimensions(pdf_path: Path) -> list[tuple[int, float, float]]:
    """Reuse exact piece dimensions for an unchanged sketch PDF."""
    cached = shower_cache.load("pdf_piece_dimensions_v1", pdf_path)
    if isinstance(cached, list):
        try:
            return [
                (int(page), float(width), float(height))
                for page, width, height in cached
            ]
        except (TypeError, ValueError):
            pass
    reader = PdfReader(str(pdf_path))
    dimensions = pdf_piece_dimensions(reader)
    shower_cache.store(
        "pdf_piece_dimensions_v1",
        pdf_path,
        [[page, width, height] for page, width, height in dimensions],
    )
    return dimensions


def dimensions_match(
    expected: tuple[float, float],
    actual: tuple[float, float],
    tolerance: float = PDF_DIMENSION_MATCH_TOLERANCE,
) -> bool:
    expected_width, expected_height = expected
    actual_width, actual_height = actual
    direct = (
        abs(expected_width - actual_width) <= tolerance
        and abs(expected_height - actual_height) <= tolerance
    )
    swapped = (
        abs(expected_width - actual_height) <= tolerance
        and abs(expected_height - actual_width) <= tolerance
    )
    return direct or swapped


def manual_dimension_match_override_enabled(
    config: dict[str, object] | None,
    aw_order: str,
) -> bool:
    """Return True only for an explicit operator-approved order override."""
    if not isinstance(config, dict):
        return False
    overrides = config.get("dimension_match_overrides", {})
    if not isinstance(overrides, dict):
        return False
    value = overrides.get(str(aw_order))
    if isinstance(value, dict):
        return bool(value.get("enabled", False))
    return bool(value)


def _record_dimension_match_note(process_order: ProcessOrder, item: int, note: str) -> None:
    notes = getattr(process_order, "dimension_match_notes", None)
    if not isinstance(notes, dict):
        notes = {}
        setattr(process_order, "dimension_match_notes", notes)
    notes[int(item)] = str(note)


def _record_manual_dimension_override(process_order: ProcessOrder) -> None:
    setattr(process_order, "manual_dimension_match_override_used", True)


def _dxf_oos_profile(
    path: Path,
    expected: tuple[float, float],
    actual: tuple[float, float] | None = None,
) -> dict[str, object] | None:
    """Describe alternate sketch dimensions proven by the source DXF.

    A+W process lists can report a different size representation than the PDF
    sketch on a parallelogram, trapezoid, or other intentionally out-of-square
    shower piece. The original fallback requires the DXF overall bounds to
    match A+W. A second, still-strict path is available when ``actual`` is
    supplied: the PDF sketch must match the source-DXF envelope, A+W must stay
    within a small process-size variance, and the DXF must prove real skew.
    """
    segments = programmer.collect_dxf_outer_line_segments(path)
    if len(segments) < 4:
        return None
    points = [point for start, end in segments for point in (start, end)]
    min_x = min(x for x, _ in points)
    max_x = max(x for x, _ in points)
    min_y = min(y for _, y in points)
    max_y = max(y for _, y in points)
    raw_width = max_x - min_x
    raw_height = max_y - min_y
    if raw_width <= 0 or raw_height <= 0:
        return None

    expected_width, expected_height = expected
    for scale in (1.0, 1.0 / 25.4):
        width = raw_width * scale
        height = raw_height * scale
        direct = (
            abs(width - expected_width) <= OOS_DXF_OVERALL_TOLERANCE
            and abs(height - expected_height) <= OOS_DXF_OVERALL_TOLERANCE
        )
        swapped = (
            abs(width - expected_height) <= OOS_DXF_OVERALL_TOLERANCE
            and abs(height - expected_width) <= OOS_DXF_OVERALL_TOLERANCE
        )
        match_basis = "aw_dxf_bounds"
        if not direct and not swapped and actual is not None:
            sketch_matches_dxf = dimensions_match(
                (width, height),
                actual,
                tolerance=OOS_DXF_OVERALL_TOLERANCE,
            )
            direct = (
                sketch_matches_dxf
                and abs(width - expected_width) <= OOS_PROCESS_DXF_VARIANCE_TOLERANCE
                and abs(height - expected_height) <= OOS_PROCESS_DXF_VARIANCE_TOLERANCE
            )
            swapped = (
                sketch_matches_dxf
                and abs(width - expected_height) <= OOS_PROCESS_DXF_VARIANCE_TOLERANCE
                and abs(height - expected_width) <= OOS_PROCESS_DXF_VARIANCE_TOLERANCE
            )
            if direct or swapped:
                match_basis = "sketch_dxf_envelope"
        if not direct and not swapped:
            continue

        # Classify long layer-0 outline segments by their dominant axis rather
        # than by closeness to a bounding-box side. This keeps strongly raked
        # trapezoid edges available as evidence even when they move several
        # inches across the piece.
        horizontal_edges: list[tuple[float, float, float]] = []
        vertical_edges: list[tuple[float, float, float]] = []
        for start_point, end_point in segments:
            dx = abs(end_point[0] - start_point[0]) * scale
            dy = abs(end_point[1] - start_point[1]) * scale
            length = math.hypot(dx, dy)
            if dx >= max(1.0, width * 0.45) and dx >= dy:
                horizontal_edges.append((dx, dy, length))
            if dy >= max(1.0, height * 0.45) and dy >= dx:
                vertical_edges.append((dx, dy, length))
        if not horizontal_edges or not vertical_edges:
            continue

        edge_width = sum(edge[0] for edge in horizontal_edges) / len(horizontal_edges)
        edge_height = sum(edge[1] for edge in vertical_edges) / len(vertical_edges)
        side_shift_x = max(edge[0] for edge in vertical_edges)
        side_shift_y = max(edge[1] for edge in horizontal_edges)
        skew_shift = max(side_shift_x, side_shift_y)
        if skew_shift < OOS_MIN_SHIFT_INCHES:
            continue

        maximum_reasonable_skew = max(2.0, max(width, height) * 0.20)
        if skew_shift > maximum_reasonable_skew:
            continue

        def unique_candidates(values: Iterable[float], minimum: float) -> list[float]:
            result: list[float] = []
            for value in values:
                value = float(value)
                if value + OOS_DXF_EDGE_TOLERANCE < minimum:
                    continue
                if not any(abs(value - existing) <= 0.01 for existing in result):
                    result.append(value)
            return result

        width_candidates = unique_candidates(
            [value for dx, _dy, length in horizontal_edges for value in (dx, length)]
            + [edge_width, width],
            max(1.0, width * 0.45),
        )
        height_candidates = unique_candidates(
            [value for _dx, dy, length in vertical_edges for value in (dy, length)]
            + [edge_height, height],
            max(1.0, height * 0.45),
        )

        profile: dict[str, object] = {
            "overall_width": width,
            "overall_height": height,
            "edge_width": edge_width,
            "edge_height": edge_height,
            "width_shift": abs(width - edge_width),
            "height_shift": abs(height - edge_height),
            "side_shift_x": side_shift_x,
            "side_shift_y": side_shift_y,
            "skew_shift": skew_shift,
            "width_candidates": width_candidates,
            "height_candidates": height_candidates,
            "match_basis": match_basis,
            "process_width_delta": abs(width - expected_width),
            "process_height_delta": abs(height - expected_height),
        }
        if swapped:
            profile = {
                "overall_width": height,
                "overall_height": width,
                "edge_width": edge_height,
                "edge_height": edge_width,
                "width_shift": abs(height - edge_height),
                "height_shift": abs(width - edge_width),
                "side_shift_x": side_shift_y,
                "side_shift_y": side_shift_x,
                "skew_shift": skew_shift,
                "width_candidates": height_candidates,
                "height_candidates": width_candidates,
                "match_basis": match_basis,
                "process_width_delta": abs(height - expected_width),
                "process_height_delta": abs(width - expected_height),
            }
        return profile
    return None


def _matched_oos_edge_pair(
    actual: tuple[float, float],
    profile: dict[str, object],
) -> tuple[float, float] | None:
    width_candidates = [
        float(value) for value in profile.get("width_candidates", [])
        if isinstance(value, (int, float))
    ]
    height_candidates = [
        float(value) for value in profile.get("height_candidates", [])
        if isinstance(value, (int, float))
    ]
    if not width_candidates:
        width_candidates = [float(profile["edge_width"])]
    if not height_candidates:
        height_candidates = [float(profile["edge_height"])]

    best: tuple[float, float, float] | None = None
    for width in width_candidates:
        for height in height_candidates:
            direct_error = abs(actual[0] - width) + abs(actual[1] - height)
            swapped_error = abs(actual[0] - height) + abs(actual[1] - width)
            if (
                abs(actual[0] - width) <= OOS_DXF_EDGE_TOLERANCE
                and abs(actual[1] - height) <= OOS_DXF_EDGE_TOLERANCE
            ):
                candidate = (direct_error, width, height)
            elif (
                abs(actual[0] - height) <= OOS_DXF_EDGE_TOLERANCE
                and abs(actual[1] - width) <= OOS_DXF_EDGE_TOLERANCE
            ):
                candidate = (swapped_error, height, width)
            else:
                continue
            if best is None or candidate[0] < best[0]:
                best = candidate
    return None if best is None else (best[1], best[2])


def _pdf_dimensions_match_oos_profile(
    actual: tuple[float, float],
    profile: dict[str, object],
) -> bool:
    return _matched_oos_edge_pair(actual, profile) is not None


def _find_oos_dxf_evidence(
    folder: Path,
    process_order: ProcessOrder,
    item_number: int,
    expected: tuple[float, float],
    actual: tuple[float, float] | None = None,
) -> tuple[Path, dict[str, object]] | None:
    panel = programmer.Panel(
        int(item_number),
        0,
        "",
        float(expected[0]),
        float(expected[1]),
        "",
    )
    source = programmer.find_source_dxf(
        folder,
        process_order.job_name,
        panel,
        process_order.aw_order,
    )
    if source is None:
        return None
    profile = _dxf_oos_profile(source, expected, actual)
    if profile is None:
        return None
    return source, profile


def reconcile_out_of_square_dimension_match(
    process_order: ProcessOrder,
    actual: list[tuple[int, float, float]],
    folder: Path | None,
) -> bool:
    """Validate an already-selected PDF using source-DXF geometry evidence.

    This fallback is intentionally not used by duplicate-Job PDF selection.
    Every process-list item must still map to a unique sketch piece; an OOS or
    raked candidate may never reuse a sketch page already needed by a strict
    dimension match.
    """
    if folder is None:
        return False
    expected = process_order_dimensions(process_order)
    if not expected or not actual or len(expected) > len(actual):
        return False

    strict_candidates: list[list[int]] = [
        [
            actual_index
            for actual_index, (_page, actual_width, actual_height) in enumerate(actual)
            if dimensions_match((width, height), (actual_width, actual_height))
        ]
        for _item, width, height in expected
    ]
    unmatched_indices = [
        expected_index
        for expected_index, candidates in enumerate(strict_candidates)
        if not candidates
    ]
    if not unmatched_indices:
        return False

    candidate_indices = [list(candidates) for candidates in strict_candidates]
    evidence_by_expected: dict[int, dict[int, tuple[Path, dict[str, object]]]] = {}
    for expected_index in unmatched_indices:
        item, width, height = expected[expected_index]
        evidence = _find_oos_dxf_evidence(folder, process_order, item, (width, height))
        matches: list[int] = []
        evidence_by_actual: dict[int, tuple[Path, dict[str, object]]] = {}
        if evidence is not None:
            source, profile = evidence
            for actual_index, (_page, actual_width, actual_height) in enumerate(actual):
                if _pdf_dimensions_match_oos_profile((actual_width, actual_height), profile):
                    matches.append(actual_index)
                    evidence_by_actual[actual_index] = (source, profile)
        else:
            # Some A+W irregular-shape process dimensions are not the raw DXF
            # bounds. In that case, evaluate each sketch candidate against the
            # source DXF itself. This path is accepted only when the sketch
            # matches the DXF envelope, A+W stays within the narrow process
            # variance, and the same DXF proves non-axis-aligned geometry.
            for actual_index, (_page, actual_width, actual_height) in enumerate(actual):
                candidate_evidence = _find_oos_dxf_evidence(
                    folder,
                    process_order,
                    item,
                    (width, height),
                    (actual_width, actual_height),
                )
                if candidate_evidence is None:
                    continue
                source, profile = candidate_evidence
                if not _pdf_dimensions_match_oos_profile((actual_width, actual_height), profile):
                    continue
                matches.append(actual_index)
                evidence_by_actual[actual_index] = (source, profile)
        if not matches:
            return False
        candidate_indices[expected_index] = matches
        evidence_by_expected[expected_index] = evidence_by_actual

    assignment: dict[int, int] = {}
    assignment_order = sorted(range(len(expected)), key=lambda index: len(candidate_indices[index]))

    def assign(position: int, used_actual: set[int]) -> bool:
        if position >= len(assignment_order):
            return True
        expected_index = assignment_order[position]
        for actual_index in candidate_indices[expected_index]:
            if actual_index in used_actual:
                continue
            assignment[expected_index] = actual_index
            used_actual.add(actual_index)
            if assign(position + 1, used_actual):
                return True
            used_actual.remove(actual_index)
            assignment.pop(expected_index, None)
        return False

    if not assign(0, set()):
        return False

    for expected_index in unmatched_indices:
        item, width, height = expected[expected_index]
        actual_index = assignment[expected_index]
        source, profile = evidence_by_expected[expected_index][actual_index]
        _page, actual_width, actual_height = actual[actual_index]
        matched_edge = _matched_oos_edge_pair((actual_width, actual_height), profile)
        edge_width, edge_height = matched_edge or (
            float(profile["edge_width"]),
            float(profile["edge_height"]),
        )
        shift = float(profile.get("skew_shift", max(
            float(profile.get("width_shift", 0.0)),
            float(profile.get("height_shift", 0.0)),
        )))
        match_basis = str(profile.get("match_basis", "aw_dxf_bounds"))
        if match_basis == "sketch_dxf_envelope":
            width_delta = float(profile.get("process_width_delta", abs(width - actual_width)))
            height_delta = float(profile.get("process_height_delta", abs(height - actual_height)))
            note = (
                f"A+W process size {width:g} x {height:g} reconciled to sketch "
                f"{actual_width:g} x {actual_height:g} because the sketch matches the "
                f"source DXF envelope ({source.name}) and the DXF proves OOS geometry "
                f"(edge {edge_width:g} x {edge_height:g}; OOS shift {shift:g} in; "
                f"A+W/DXF delta {width_delta:g} x {height_delta:g} in)."
            )
        else:
            note = (
                f"A+W overall size {width:g} x {height:g} reconciled to sketch "
                f"{actual_width:g} x {actual_height:g} by DXF OOS geometry "
                f"({source.name}; edge {edge_width:g} x {edge_height:g}; "
                f"OOS shift {shift:g} in)."
            )
        _record_dimension_match_note(process_order, item, note)
    return True


def process_dimensions_fit_pdf(
    process_order: ProcessOrder,
    reader: PdfReader,
) -> bool | None:
    return process_dimensions_fit_values(process_order, pdf_piece_dimensions(reader))


def process_dimensions_fit_values(
    process_order: ProcessOrder,
    actual: list[tuple[int, float, float]],
) -> bool | None:
    """Apply the existing PDF dimension rules to pre-extracted dimensions."""
    expected = process_order_dimensions(process_order)
    if not expected or not actual:
        return None
    if len(expected) > len(actual):
        return False

    candidates = [
        [
            actual_index
            for actual_index, (_page, actual_width, actual_height) in enumerate(actual)
            if dimensions_match((width, height), (actual_width, actual_height))
        ]
        for _item, width, height in expected
    ]
    if any(not matches for matches in candidates):
        return False

    order = sorted(range(len(expected)), key=lambda index: len(candidates[index]))

    def assign(position: int, used: set[int]) -> bool:
        if position >= len(order):
            return True
        expected_index = order[position]
        for actual_index in candidates[expected_index]:
            if actual_index in used:
                continue
            used.add(actual_index)
            if assign(position + 1, used):
                return True
            used.remove(actual_index)
        return False

    return assign(0, set())


def format_dimension_pairs(dimensions: Iterable[tuple[int, float, float]]) -> str:
    return ", ".join(f"P{item} {width:g} x {height:g}" for item, width, height in dimensions)


def dimension_mismatch_difference_lines(
    process_order: ProcessOrder,
    actual_values: list[tuple[int, float, float]],
) -> list[str]:
    """Explain the closest process-list/sketch numeric difference per item."""
    lines: list[str] = []
    for item, expected_width, expected_height in process_order_dimensions(process_order):
        best: tuple[float, int, float, float, float, float] | None = None
        for page, actual_width, actual_height in actual_values:
            direct = (
                abs(expected_width - actual_width),
                abs(expected_height - actual_height),
                actual_width,
                actual_height,
            )
            swapped = (
                abs(expected_width - actual_height),
                abs(expected_height - actual_width),
                actual_height,
                actual_width,
            )
            for width_delta, height_delta, matched_width, matched_height in (direct, swapped):
                score = max(width_delta, height_delta) + width_delta + height_delta
                candidate = (score, page, width_delta, height_delta, matched_width, matched_height)
                if best is None or candidate[0] < best[0]:
                    best = candidate
        if best is None:
            continue
        _score, page, width_delta, height_delta, _matched_width, _matched_height = best
        over_width = max(0.0, width_delta - PDF_DIMENSION_MATCH_TOLERANCE)
        over_height = max(0.0, height_delta - PDF_DIMENSION_MATCH_TOLERANCE)
        detail = (
            f"  P{item} vs PDF page {page + 1}: width difference {width_delta:.4f} in; "
            f"height difference {height_delta:.4f} in; normal tolerance {PDF_DIMENSION_MATCH_TOLERANCE:.4f} in"
        )
        exceeded: list[str] = []
        if over_width > 0:
            exceeded.append(f"width exceeds tolerance by {over_width:.4f} in")
        if over_height > 0:
            exceeded.append(f"height exceeds tolerance by {over_height:.4f} in")
        if exceeded:
            detail += " (" + "; ".join(exceeded) + ")"
        lines.append(detail)
    return lines


def dimension_mismatch_message(
    process_order: ProcessOrder,
    pdf_path: Path,
    actual_values: list[tuple[int, float, float]],
) -> str:
    """Build a readable operator-facing dimension mismatch explanation."""
    expected_lines = [
        f"  P{item}: {width:g} x {height:g}"
        for item, width, height in process_order_dimensions(process_order)
    ] or ["  Unknown"]
    sketch_lines = [
        f"  PDF page {page + 1}: {width:g} x {height:g}"
        for page, width, height in actual_values
    ] or ["  Unknown"]
    difference_lines = dimension_mismatch_difference_lines(process_order, actual_values)
    return (
        f"A&W {process_order.aw_order} does not match the piece dimensions in:\n"
        f"{pdf_path.name}\n\n"
        "PROCESS LIST\n"
        + "\n".join(expected_lines)
        + "\n\nSKETCH\n"
        + "\n".join(sketch_lines)
        + ("\n\nNUMERIC DIFFERENCE\n" + "\n".join(difference_lines) if difference_lines else "")
        + "\n\nWHAT THE PROGRAMMER CHECKED\n"
        "The normal dimension match failed. The programmer also checked the matching DXF for "
        "out-of-square, raked, or other alternate edge geometry, but could not prove that the two "
        "dimension sets describe the same piece.\n\n"
        "NEXT STEP\n"
        "Verify the sketch and DXF. If they are intentionally correct, right-click the order and choose "
        "Allow Dimension Mismatch."
    )


def validate_process_order_pdf_dimensions(
    reader: PdfReader,
    process_order: ProcessOrder,
    pdf_path: Path,
    *,
    folder: Path | None = None,
    config: dict[str, object] | None = None,
) -> None:
    actual_values = pdf_piece_dimensions(reader)
    fit = process_dimensions_fit_values(process_order, actual_values)
    if fit is not False:
        return
    if reconcile_out_of_square_dimension_match(process_order, actual_values, folder):
        return
    if manual_dimension_match_override_enabled(config, process_order.aw_order):
        _record_manual_dimension_override(process_order)
        return
    raise shower_errors.ShowerProgrammerError(
        code=shower_errors.ErrorCode.DIMENSION_MISMATCH,
        title="Dimension mismatch",
        message=dimension_mismatch_message(process_order, pdf_path, actual_values),
        aw_order=str(process_order.aw_order),
        details={"pdf": str(pdf_path)},
    )


def validate_process_order_pdf_dimension_values(
    actual: list[tuple[int, float, float]],
    process_order: ProcessOrder,
    pdf_path: Path,
    *,
    folder: Path | None = None,
    config: dict[str, object] | None = None,
) -> None:
    fit = process_dimensions_fit_values(process_order, actual)
    if fit is not False:
        return
    if reconcile_out_of_square_dimension_match(process_order, actual, folder):
        return
    if manual_dimension_match_override_enabled(config, process_order.aw_order):
        _record_manual_dimension_override(process_order)
        return
    raise shower_errors.ShowerProgrammerError(
        code=shower_errors.ErrorCode.DIMENSION_MISMATCH,
        title="Dimension mismatch",
        message=dimension_mismatch_message(process_order, pdf_path, actual),
        aw_order=str(process_order.aw_order),
        details={"pdf": str(pdf_path)},
    )


def job_number_pdf_candidates(
    folder: Path,
    process_order: ProcessOrder,
    candidate_pdfs: Iterable[Path] | None = None,
) -> list[Path]:
    job_number = programmer.extract_job_number(process_order.job_name)
    if not job_number:
        return []
    candidates = folder.rglob("*.pdf") if candidate_pdfs is None else candidate_pdfs
    return sorted(
        (
            Path(path)
            for path in candidates
            if Path(path).suffix.lower() == ".pdf"
            and not programmer.is_archived_input_file(Path(path), folder)
            and programmer.text_contains_job_number(Path(path).stem, job_number)
        ),
        key=lambda path: path.name.casefold(),
    )


def dimension_matched_pdf(
    folder: Path,
    process_order: ProcessOrder,
    *,
    exclude: Path | None = None,
    candidate_pdfs: Iterable[Path] | None = None,
) -> tuple[Path, PdfReader] | None:
    matches: list[tuple[Path, PdfReader]] = []
    for candidate in job_number_pdf_candidates(folder, process_order, candidate_pdfs):
        if exclude is not None and candidate.resolve() == exclude.resolve():
            continue
        try:
            reader = PdfReader(str(candidate))
        except Exception:
            continue
        if process_dimensions_fit_pdf(process_order, reader) is True:
            matches.append((candidate.resolve(), reader))
    if len(matches) == 1:
        return matches[0]
    return None


def dimension_matched_pdf_path(
    folder: Path,
    process_order: ProcessOrder,
    *,
    exclude: Path | None = None,
    candidate_pdfs: Iterable[Path] | None = None,
) -> Path | None:
    """Resolve duplicate Job Nr sketches using cached dimension evidence."""
    matches: list[Path] = []
    for candidate in job_number_pdf_candidates(folder, process_order, candidate_pdfs):
        if exclude is not None and candidate.resolve() == exclude.resolve():
            continue
        try:
            actual = cached_pdf_piece_dimensions(candidate)
        except Exception:
            continue
        if process_dimensions_fit_values(process_order, actual) is True:
            matches.append(candidate.resolve())
    return matches[0] if len(matches) == 1 else None


def preview_process_order_pdf(
    folder: Path,
    process_order: ProcessOrder,
    candidate_pdfs: list[Path],
    config: dict[str, object] | None = None,
) -> Path:
    """Resolve and validate a scan preview without reparsing unchanged PDFs."""
    try:
        pdf_path = programmer.find_pdf(
            folder,
            process_order.job_name,
            process_order.aw_order,
            candidate_pdfs=candidate_pdfs,
        ).resolve()
    except RuntimeError:
        matched = dimension_matched_pdf_path(
            folder,
            process_order,
            candidate_pdfs=candidate_pdfs,
        )
        if matched is None:
            raise
        pdf_path = matched
        validate_process_order_pdf_dimension_values(
            cached_pdf_piece_dimensions(pdf_path),
            process_order,
            pdf_path,
            folder=folder,
            config=config,
        )
        return pdf_path

    try:
        validate_process_order_pdf_dimension_values(
            cached_pdf_piece_dimensions(pdf_path),
            process_order,
            pdf_path,
            folder=folder,
            config=config,
        )
    except RuntimeError:
        matched = dimension_matched_pdf_path(
            folder,
            process_order,
            exclude=pdf_path,
            candidate_pdfs=candidate_pdfs,
        )
        if matched is None:
            raise
        pdf_path = matched
        validate_process_order_pdf_dimension_values(
            cached_pdf_piece_dimensions(pdf_path),
            process_order,
            pdf_path,
            folder=folder,
            config=config,
        )
    return pdf_path


def open_process_order_pdf(
    folder: Path,
    process_order: ProcessOrder,
    config: dict[str, object] | None = None,
) -> tuple[Path, PdfReader]:
    try:
        pdf_path = programmer.find_pdf(folder, process_order.job_name, process_order.aw_order).resolve()
    except RuntimeError:
        matched = dimension_matched_pdf(folder, process_order)
        if matched is None:
            raise
        pdf_path, reader = matched
        validate_process_order_pdf_dimensions(
            reader,
            process_order,
            pdf_path,
            folder=folder,
            config=config,
        )
        return pdf_path, reader

    reader = PdfReader(str(pdf_path))
    try:
        validate_process_order_pdf_dimensions(
            reader,
            process_order,
            pdf_path,
            folder=folder,
            config=config,
        )
    except RuntimeError:
        matched = dimension_matched_pdf(folder, process_order, exclude=pdf_path)
        if matched is None:
            raise
        pdf_path, reader = matched
        validate_process_order_pdf_dimensions(
            reader,
            process_order,
            pdf_path,
            folder=folder,
            config=config,
        )
    return pdf_path, reader


LOCATION_FIELD_RE = re.compile(
    r"location\s*:\s*(.*?)"
    r"(?="
    r"\s*(?:marks?|project(?:\s*#)?|shape|quantity|glass|page)\s*:"
    r"|\s+measurements\s+are\s+in\s+inches"
    r"|\s+page\s+\d+\s+of\s+\d+"
    r"|$"
    r")",
    re.IGNORECASE,
)


def pdf_location_values(reader: PdfReader) -> list[str]:
    """Return A+W Location values from overview and piece-page content.

    A+W PDF text extraction sometimes joins the preceding value directly to
    ``Location:`` (for example ``MASTERLocation:``) or places the location
    value on the next visual line.  Collapsing page whitespace before applying
    the field-boundary expression handles both forms without treating an
    unrelated REMAKE note elsewhere on the page as the Location value.
    """
    values: list[str] = []
    for page in reader.pages:
        text = re.sub(r"\s+", " ", page.extract_text() or "").strip()
        for match in LOCATION_FIELD_RE.finditer(text):
            value = re.sub(r"\s+", " ", match.group(1)).strip(" :-")
            if value:
                values.append(value)
    return values


def pdf_location_indicates_remake(reader: PdfReader) -> bool:
    """Detect REMAKE whenever the A+W Location field says REMAKE."""
    return any(re.search(r"\bREMAKE\b", value, re.IGNORECASE) for value in pdf_location_values(reader))


def prepare_job(
    folder: Path,
    sketch_output_dir: Path,
    dxf_output_dir: Path,
    report_dir: Path,
    config: dict[str, object],
    process_order: ProcessOrder,
    remake_items: set[int] | None = None,
) -> tuple[programmer.Job, PdfReader, list[str]]:
    process_order = clone_process_order(process_order)
    pdf_path, reader = open_process_order_pdf(folder, process_order, config=config)
    if remake_items is None and pdf_location_indicates_remake(reader):
        remake_items = set()
    panels = programmer.analyze_panels(reader, config, process_order.aw_order)
    item_remaps = match_process_items_to_sketch_pages(reader, panels, process_order, config, remake_items)
    if remake_items:
        remake_items = {item_remaps.get(item, item) for item in remake_items}
    attach_unlabeled_process_pages(reader, panels, process_order, config)
    attach_unmarked_process_pages(reader, panels, process_order, config)
    reconcile_process_list_item_gaps(panels, process_order)
    reconcile_missing_items_from_extra_sketch_pages(panels, process_order)
    apply_process_hints(panels, process_order, config)
    apply_process_list_scope(panels, process_order)
    programmer.refine_panel_orientations(reader, panels, config)
    for panel in panels:
        programmer.apply_override(panel, config, process_order.aw_order)
    apply_process_list_scope(panels, process_order)
    effective_remake_items = (
        set(process_order.item_numbers)
        if remake_items is not None and not remake_items
        else remake_items
    )
    selected_remake_items = programmer.apply_remake_selection(panels, effective_remake_items)

    job = programmer.Job(
        pdf_path=pdf_path,
        aw_order=process_order.aw_order,
        job_name=process_order.job_name,
        panels=panels,
        output_pdf=sketch_output_dir / f"{process_order.aw_order}.pdf",
        report_path=report_dir / f"{process_order.aw_order}_programming_report.txt",
        remake_items=selected_remake_items,
    )
    programmer.assign_dxf_paths(job, folder, dxf_output_dir, config)
    programmer.apply_corner_text_indicator_avoidance(reader, job.panels, config)
    return job, reader, collect_issues(job, process_order)


def collect_issues(job: programmer.Job, process_order: ProcessOrder) -> list[str]:
    issues: list[str] = []
    expected = set(process_order.item_numbers)
    actual = {panel.item for panel in job.panels}
    missing = sorted(expected - actual)
    extra = sorted(actual - expected)
    if missing:
        issues.append("Missing sketch page: " + ", ".join(f"P{i}" for i in missing))
    if extra:
        issues.append("Sketch page not in list; crossed out: " + ", ".join(f"P{i}" for i in extra))

    for panel in job.panels:
        if panel.remake_excluded:
            continue
        has_process_remap_reason = any(
            reason.startswith("process-list P") and "applied to sketch" in reason
            for reason in panel.reasons
        )
        if panel.source_item is not None and panel.source_item != panel.item and not has_process_remap_reason:
            issues.append(f"P{panel.item}: source item P{panel.source_item}")
        for reason in panel.reasons:
            if reason.startswith("process-list P") and "applied to sketch" in reason:
                issues.append(f"P{panel.item}: {reason}")
        for warning in panel.warnings:
            issues.append(f"P{panel.item}: {concise_issue_message(warning)}")
        if not panel.skip_dxf and panel.source_dxf is None:
            issues.append(f"P{panel.item}: missing DXF")
    return list(dict.fromkeys(issues))


def concise_issue_message(message: str) -> str:
    normalized = re.sub(r"\s+", " ", message).strip()
    replacements = {
        "No matching process-list item row found.": "no process-list row",
        "No matching source DXF found.": "missing DXF",
        "Could not read dimensions from PDF text.": "could not read size",
        "Label placed in best available open area inside glass.": "label placed inside best open area",
        "FP-S cut-in/cut-out detected; manual DXF review required.": "FP-S cut; review DXF",
    }
    return replacements.get(normalized, normalized)


def process_one_order(
    process_order: ProcessOrder,
    folder: Path,
    sketch_output_dir: Path,
    dxf_output_dir: Path,
    report_dir: Path,
    config: dict[str, object],
    apply: bool,
    force: bool,
    skip_pdf: bool,
    skip_dxf: bool,
    remake_items: set[int] | None = None,
) -> BatchJobResult:
    item_text = ", ".join(f"P{i}" for i in process_order.item_numbers)
    result = BatchJobResult(
        aw_order=process_order.aw_order,
        job_name=process_order.job_name,
        customer=process_order.customer,
        items=item_text,
        status="PENDING",
        delivery_date=process_order.delivery_date,
    )

    try:
        job, reader, issues = prepare_job(
            folder,
            sketch_output_dir,
            dxf_output_dir,
            report_dir,
            config,
            process_order,
            remake_items=remake_items,
        )
        programmed_items = sorted({panel.item for panel in job.panels if not panel.remake_excluded})
        if programmed_items:
            result.items = ", ".join(f"P{i}" for i in programmed_items)
        result.input_pdf = job.pdf_path
        result.output_pdf = job.output_pdf
        result.report_path = job.report_path
        result.remake_items = None if job.remake_items is None else sorted(job.remake_items)
        result.issues.extend(issues)
        if apply:
            if not force and not skip_pdf and job.output_pdf.exists():
                result.status = "SKIPPED"
                result.issues.append(f"Output PDF already exists: {job.output_pdf.name}")
                return result
            if not skip_pdf:
                programmer.write_marked_pdf(job, reader, config, force=force)
            if not skip_dxf:
                result.issues.extend(write_dxfs_with_issue_collection(job, force=force, config=config))
            write_individual_report(job, result.issues, apply, skip_pdf, skip_dxf, force=True)

        result.status = "ISSUES" if result.issues else "OK"
        return result
    except Exception as exc:
        result.status = "FAILED"
        result.issues.append(str(exc))
        return result


def write_dxfs_with_issue_collection(job: programmer.Job, force: bool, config: dict[str, object]) -> list[str]:
    issues: list[str] = []
    for panel in job.panels:
        if panel.skip_dxf:
            continue
        if panel.source_dxf is None:
            issues.append(f"P{panel.item}: source DXF missing, skipped")
            continue
        if panel.output_dxf is None:
            issues.append(f"P{panel.item}: output DXF path missing, skipped")
            continue
        try:
            programmer.write_panel_dxf(panel, force=force, config=config)
        except Exception as exc:
            issues.append(f"P{panel.item}: DXF failed, {exc}")
    return issues


def write_individual_report(
    job: programmer.Job,
    issues: list[str],
    apply: bool,
    skip_pdf: bool,
    skip_dxf: bool,
    force: bool,
) -> None:
    report = programmer.build_report(job, apply=apply, skip_pdf=skip_pdf, skip_dxf=skip_dxf)
    if issues:
        report += "\nBatch issues:\n"
        report += "\n".join(f"- {issue}" for issue in issues) + "\n"
    programmer.write_report(job.report_path, report, force=force)


def run_batch(
    orders: list[ProcessOrder],
    folder: Path,
    sketch_output_dir: Path,
    dxf_output_dir: Path,
    report_dir: Path,
    config: dict[str, object],
    apply: bool,
    force: bool,
    skip_pdf: bool = False,
    skip_dxf: bool = False,
    remake_items_by_order: dict[str, set[int]] | None = None,
    progress: Callable[[BatchJobResult], None] | None = None,
) -> BatchRunResult:
    sketch_output_dir.mkdir(parents=True, exist_ok=True)
    dxf_output_dir.mkdir(parents=True, exist_ok=True)
    report_dir.mkdir(parents=True, exist_ok=True)

    results: list[BatchJobResult] = []
    for process_order in orders:
        result = process_one_order(
            process_order=process_order,
            folder=folder,
            sketch_output_dir=sketch_output_dir,
            dxf_output_dir=dxf_output_dir,
            report_dir=report_dir,
            config=config,
            apply=apply,
            force=force,
            skip_pdf=skip_pdf,
            skip_dxf=skip_dxf,
            remake_items=None if remake_items_by_order is None else remake_items_by_order.get(process_order.aw_order),
        )
        results.append(result)
        if progress:
            progress(result)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    text_report = report_dir / f"batch_programming_report_{timestamp}.txt"
    csv_report = report_dir / f"batch_programming_report_{timestamp}.csv"
    html_report = report_dir / f"batch_programming_report_{timestamp}.html"
    write_batch_text_report(text_report, results, apply)
    write_batch_csv_report(csv_report, results)
    write_batch_html_report(html_report, results, apply)
    return BatchRunResult(results=results, text_report=text_report, csv_report=csv_report, html_report=html_report)


def write_batch_text_report(path: Path, results: list[BatchJobResult], apply: bool) -> None:
    counts = count_statuses(results)
    lines = [
        f"Shower batch {'APPLY' if apply else 'DRY RUN'} report",
        f"Created: {datetime.now():%Y-%m-%d %H:%M:%S}",
        f"Total orders: {len(results)}",
        "Status counts: " + ", ".join(f"{status}={count}" for status, count in counts.items()),
        "",
    ]
    for result in results:
        lines.append(f"{result.status}: {result.aw_order} | {result.job_name} | {result.items}")
        if result.delivery_date:
            lines.append(f"  delivery date: {result.delivery_date}")
        if result.input_pdf:
            lines.append(f"  input pdf: {result.input_pdf}")
        if result.output_pdf:
            lines.append(f"  output pdf: {result.output_pdf}")
        if result.report_path:
            lines.append(f"  order report: {result.report_path}")
        for issue in result.issues:
            lines.append(f"  issue: {issue}")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_batch_csv_report(path: Path, results: list[BatchJobResult]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Status", "A&W Order", "Delivery Date", "Job", "Customer", "Items", "Input PDF", "Output PDF", "Issues"])
        for result in results:
            writer.writerow(
                [
                    result.status,
                    result.aw_order,
                    result.delivery_date,
                    result.job_name,
                    result.customer,
                    result.items,
                    str(result.input_pdf or ""),
                    str(result.output_pdf or ""),
                    " | ".join(result.issues),
                ]
            )


def write_batch_html_report(path: Path, results: list[BatchJobResult], apply: bool) -> None:
    counts = count_statuses(results)
    rows = []
    for result in results:
        rows.append(
            "<tr>"
            f"<td class='{html.escape(result.status.lower())}'>{html.escape(result.status)}</td>"
            f"<td>{html.escape(result.aw_order)}</td>"
            f"<td>{html.escape(result.delivery_date)}</td>"
            f"<td>{html.escape(result.job_name)}</td>"
            f"<td>{html.escape(result.customer)}</td>"
            f"<td>{html.escape(result.items)}</td>"
            f"<td>{path_link(result.input_pdf)}</td>"
            f"<td>{path_link(result.output_pdf)}</td>"
            f"<td>{html.escape('; '.join(result.issues))}</td>"
            "</tr>"
        )
    document = f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>Shower Batch Report</title>
<style>
body {{ font-family: Arial, sans-serif; margin: 24px; color: #1f2933; }}
h1 {{ margin-bottom: 4px; }}
.summary {{ margin: 10px 0 18px; }}
table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
th, td {{ border: 1px solid #d7dde5; padding: 7px 9px; vertical-align: top; }}
th {{ background: #eef3f8; text-align: left; position: sticky; top: 0; }}
tr:nth-child(even) {{ background: #f8fafc; }}
.ok {{ color: #0f7a3b; font-weight: bold; }}
.issues {{ color: #9a5b00; font-weight: bold; }}
.failed, .skipped {{ color: #b42318; font-weight: bold; }}
</style>
</head>
<body>
<h1>Shower Batch {'Apply' if apply else 'Dry Run'} Report</h1>
<div class="summary">Created {datetime.now():%Y-%m-%d %H:%M:%S} | Total orders: {len(results)} | {html.escape(', '.join(f'{k}={v}' for k, v in counts.items()))}</div>
<table>
<thead><tr><th>Status</th><th>A&amp;W Order</th><th>Delivery Date</th><th>Job</th><th>Customer</th><th>Items</th><th>Input PDF</th><th>Output PDF</th><th>Issues</th></tr></thead>
<tbody>
{''.join(rows)}
</tbody>
</table>
</body>
</html>
"""
    path.write_text(document, encoding="utf-8")


def path_link(path: Path | None) -> str:
    if path is None:
        return ""
    label = html.escape(path.name)
    try:
        return f'<a href="{html.escape(path.resolve().as_uri())}">{label}</a>'
    except ValueError:
        return html.escape(str(path))


def count_statuses(results: list[BatchJobResult]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for result in results:
        counts[result.status] = counts.get(result.status, 0) + 1
    return counts


def preview_orders(
    orders: list[ProcessOrder],
    folder: Path,
    config: dict[str, object] | None = None,
) -> list[BatchJobResult]:
    candidate_pdfs = sorted(
        (
            path
            for path in folder.rglob("*.pdf")
            if path.is_file() and not programmer.is_archived_input_file(path, folder)
        ),
        key=lambda path: path.name.casefold(),
    )
    results: list[BatchJobResult] = []
    for order in orders:
        result = BatchJobResult(
            aw_order=order.aw_order,
            job_name=order.job_name,
            customer=order.customer,
            items=", ".join(f"P{i}" for i in order.item_numbers),
            status="READY",
            delivery_date=order.delivery_date,
        )
        try:
            result.input_pdf = preview_process_order_pdf(folder, order, candidate_pdfs, config=config)
        except Exception as exc:
            result.status = "ISSUES"
            result.issues.append(str(exc))
        results.append(result)
    return results


def resolve_process_list_path(process_list_value: str, folder: Path) -> Path:
    process_list = Path(process_list_value)
    if process_list.is_absolute():
        return process_list
    candidates = [
        Path.cwd() / process_list,
        programmer.project_root() / process_list,
        folder / process_list,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()
    return (folder / process_list).resolve()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Batch process shower orders from Process List Per Machine.xlsx.")
    parser.add_argument("--folder", default=str(programmer.default_orders_dir()), help="Folder containing the PDFs and DXFs.")
    parser.add_argument(
        "--process-list",
        default=str(programmer.default_process_list_path()),
        help="Path to one process-list workbook or a folder of immediate .xlsx process lists.",
    )
    parser.add_argument("--output-dir", default=str(programmer.default_output_dir()), help="Root folder for Sketches, Programs, and Reports.")
    parser.add_argument("--sketch-dir", help="Optional folder for marked PDFs. Defaults to OUTPUT\\Sketches.")
    parser.add_argument("--programs-dir", help="Optional folder for adjusted DXFs. Defaults to OUTPUT\\Programs.")
    parser.add_argument("--report-dir", help="Optional folder for reports. Defaults to OUTPUT\\Reports.")
    parser.add_argument("--dxf-output-dir", help="Optional separate folder for adjusted DXFs.")
    parser.add_argument("--config", default=programmer.DEFAULT_CONFIG_NAME, help="JSON config file.")
    parser.add_argument("--orders", help="Comma-separated A&W order numbers to process.")
    parser.add_argument("--apply", action="store_true", help="Write marked PDFs and DXFs. Without this, writes reports only.")
    parser.add_argument("--force", action="store_true", help="Overwrite existing per-order outputs.")
    parser.add_argument("--skip-pdf", action="store_true", help="Do not write marked PDFs.")
    parser.add_argument("--skip-dxf", action="store_true", help="Do not write adjusted DXFs.")
    parser.add_argument("--remake-orders", help="Comma-separated A&W order numbers to mark REMAKE.")
    parser.add_argument("--remake-items", help="Optional comma list of remake items applied to remake orders, e.g. 1,3.")
    parser.add_argument("--preview", action="store_true", help="Only list orders and PDF availability.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    folder = Path(args.folder).resolve()
    process_list = resolve_process_list_path(args.process_list, folder)
    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = programmer.project_root() / output_dir
    shower_cache.configure(output_dir.resolve() / ".scan_cache")
    shower_cache.reset_stats()
    sketch_output_dir = Path(args.sketch_dir) if args.sketch_dir else output_dir / "Sketches"
    if not sketch_output_dir.is_absolute():
        sketch_output_dir = programmer.project_root() / sketch_output_dir
    dxf_output_dir = Path(args.dxf_output_dir or args.programs_dir) if (args.dxf_output_dir or args.programs_dir) else output_dir / "Programs"
    if not dxf_output_dir.is_absolute():
        dxf_output_dir = programmer.project_root() / dxf_output_dir
    report_dir = Path(args.report_dir) if args.report_dir else output_dir / "Reports"
    if not report_dir.is_absolute():
        report_dir = programmer.project_root() / report_dir
    config = programmer.load_config(programmer.resolve_config_path(args.config, folder))
    manual_overrides = output_dir / "manual_overrides.json"
    if manual_overrides.exists():
        config = programmer.merge_item_overrides(config, programmer.load_config(manual_overrides))
    loaded_orders = load_process_orders(process_list)
    orders = filter_orders(loaded_orders, args.orders) if args.orders else visible_orders(loaded_orders, config)
    if not orders:
        raise RuntimeError("No orders found in the process list.")

    remake_items_by_order = None
    if args.remake_orders:
        remake_orders = {part.strip() for part in args.remake_orders.split(",") if part.strip()}
        remake_items = programmer.parse_item_list(args.remake_items)
        remake_items_by_order = {order.aw_order: set(remake_items) for order in orders if order.aw_order in remake_orders}

    if args.preview:
        for result in preview_orders(orders, folder, config=config):
            issue_text = "; ".join(result.issues)
            print(f"{result.status:7} {result.aw_order} {result.job_name} {result.items} {issue_text}")
        return 0

    run = run_batch(
        orders=orders,
        folder=folder,
        sketch_output_dir=sketch_output_dir,
        dxf_output_dir=dxf_output_dir,
        report_dir=report_dir,
        config=config,
        apply=args.apply,
        force=args.force,
        skip_pdf=args.skip_pdf,
        skip_dxf=args.skip_dxf,
        remake_items_by_order=remake_items_by_order,
        progress=lambda result: print(f"{result.status:7} {result.aw_order} {result.job_name}"),
    )
    print(f"\nText report: {run.text_report}")
    print(f"CSV report: {run.csv_report}")
    print(f"HTML report: {run.html_report}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
